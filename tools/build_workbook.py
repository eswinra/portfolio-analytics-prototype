"""Build outputs/Portfolio_Analytics_Dashboard_Workbook_Prototype.xlsx (stage 3).

Deterministic synthetic data (seed 20260630). Conventions:
  - blue font = hardcoded input, black = same-sheet formula, green = cross-sheet link
  - no hardcodes inside calculation formulas; assumptions live in visible cells
Also writes expected_values.json with an independent Python recomputation for QA.
"""
import datetime as dt
import json
import os
import random

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "outputs", "Portfolio_Analytics_Dashboard_Workbook_Prototype.xlsx")
QA_JSON = os.path.join(ROOT, "outputs", "expected_values.json")

# ---------------------------------------------------------------- styles
F_TITLE = Font(bold=True, size=14, color="1F3864")
F_H2 = Font(bold=True, size=11, color="1F3864")
F_HDR = Font(bold=True, size=10, color="FFFFFF")
F_INPUT = Font(color="0000CC", size=10)
F_FORMULA = Font(color="000000", size=10)
F_LINK = Font(color="006100", size=10)
F_NOTE = Font(italic=True, size=9, color="595959")
F_WARN = Font(bold=True, size=10, color="9C0006")
FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_INPUT = PatternFill("solid", fgColor="DDEBF7")
FILL_CALC = PatternFill("solid", fgColor="FFFFFF")
FILL_BAND = PatternFill("solid", fgColor="F2F2F2")
FILL_DISC = PatternFill("solid", fgColor="FFF2CC")
FILL_PASS = PatternFill("solid", fgColor="C6EFCE")
FILL_WARNF = PatternFill("solid", fgColor="FFEB9C")
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

PCT2 = "0.00%"
PCT1 = "0.0%"
MM0 = "#,##0"
MM1 = "#,##0.0"
PX4 = "0.0000"
DATE = "yyyy-mm-dd"
MON = "mmm yyyy"

DISCLAIMER = ("PROTOTYPE — SYNTHETIC / PUBLIC DATA ONLY. This workbook is an exploratory "
              "prototype for discussion. It is NOT an official LACERA system, report, or "
              "statement of performance. All DEMOFUND values are synthetic.")

# ---------------------------------------------------------------- synthetic data
rng = random.Random(20260630)
AS_OF = dt.date(2026, 6, 30)
MONTHS = []  # month-end dates Jul 2025 .. Jun 2026
d = dt.date(2025, 7, 31)
for i in range(12):
    y = 2025 + (7 + i - 1) // 12
    m = (7 + i - 1) % 12 + 1
    if m == 12:
        MONTHS.append(dt.date(y, 12, 31))
    else:
        nxt = dt.date(y + (1 if m == 12 else 0), (m % 12) + 1, 1)
        MONTHS.append(nxt - dt.timedelta(days=1))

CATS = [
    ("GROWTH", "Growth"),
    ("CREDIT", "Credit"),
    ("RAIH", "Real Assets & Inflation Hedges"),
    ("RRM", "Risk Reduction & Mitigation"),
    ("OVERLAY", "Overlays & Hedges"),
    ("OTHER", "Other Asset"),
]
CAT_IDS = [c[0] for c in CATS]

# monthly net returns per category (decimals), seeded + clamped, deliberately
# NOT matching any published LACERA figure
def gen_series(mean, vol, lo, hi):
    return [round(max(lo, min(hi, rng.gauss(mean, vol))), 4) for _ in range(12)]

PORT_R = {
    "GROWTH": gen_series(0.009, 0.028, -0.06, 0.06),
    "CREDIT": gen_series(0.006, 0.009, -0.02, 0.03),
    "RAIH": gen_series(0.005, 0.014, -0.03, 0.04),
    "RRM": gen_series(0.003, 0.008, -0.02, 0.02),
    "OVERLAY": gen_series(0.002, 0.004, -0.01, 0.01),
    "OTHER": [0.0] * 12,
}
BENCH_R = {
    cid: [round(r + rng.gauss(0, 0.004), 4) for r in PORT_R[cid]]
    for cid in ("GROWTH", "CREDIT", "RAIH", "RRM")
}

BMV0 = {"GROWTH": 4700.0, "CREDIT": 1400.0, "RAIH": 1500.0, "RRM": 2350.0,
        "OVERLAY": 40.0, "OTHER": 10.0}
# end-of-month internal transfers ($mm, net zero across categories) at quarter ends
TRANSFERS = {i: {c: 0.0 for c in CAT_IDS} for i in range(12)}
TRANSFERS[2] = {"GROWTH": -60.0, "CREDIT": 20.0, "RAIH": 15.0, "RRM": 25.0, "OVERLAY": 0.0, "OTHER": 0.0}
TRANSFERS[5] = {"GROWTH": 40.0, "CREDIT": -25.0, "RAIH": -20.0, "RRM": 5.0, "OVERLAY": 0.0, "OTHER": 0.0}
TRANSFERS[8] = {"GROWTH": -30.0, "CREDIT": 10.0, "RAIH": 25.0, "RRM": -5.0, "OVERLAY": 0.0, "OTHER": 0.0}

# policy target versions (weights %, four benchmark categories; OVERLAY/OTHER = 0)
POLICY = [
    {"effective": dt.date(2025, 7, 1), "GROWTH": 0.47, "CREDIT": 0.14, "RAIH": 0.15, "RRM": 0.24},
    {"effective": dt.date(2026, 1, 1), "GROWTH": 0.48, "CREDIT": 0.13, "RAIH": 0.15, "RRM": 0.24},
]
RANGES = {"GROWTH": 0.05, "CREDIT": 0.03, "RAIH": 0.03, "RRM": 0.04}
HURDLE_ANNUAL = 0.0675  # synthetic actuarial-style hurdle assumption

# market strip: 6 synthetic proxies x business days of June 2026
PROXIES = [
    ("DEMO-EQ-GLOBAL", "Demo global equity proxy", "Growth", 100.0),
    ("DEMO-EQ-SMALL", "Demo small-cap equity proxy", "Growth", 80.0),
    ("DEMO-BOND-AGG", "Demo aggregate bond proxy", "Risk Reduction & Mitigation", 95.0),
    ("DEMO-GOLD", "Demo gold proxy", "Real Assets & Inflation Hedges", 120.0),
    ("DEMO-OIL", "Demo crude-oil proxy", "Real Assets & Inflation Hedges", 60.0),
    ("DEMO-USD", "Demo U.S. dollar proxy", "Cross-portfolio / FX", 28.0),
]
JUN_DAYS = []
d = dt.date(2026, 6, 1)
while d <= dt.date(2026, 6, 30):
    if d.weekday() < 5:
        JUN_DAYS.append(d)
    d += dt.timedelta(days=1)
CLOSES = {}
for pid, _name, _cat, p0 in PROXIES:
    px, series = p0, []
    for _day in JUN_DAYS:
        px = round(px * (1 + rng.gauss(0.0004, 0.008)), 4)
        series.append(px)
    CLOSES[pid] = series
# deliberate data-quality demos:
CLOSES["DEMO-OIL"][-1] = None            # missing final close
STALE_PROXY = "DEMO-USD"                  # stale: last update 2 business days early
CLOSES[STALE_PROXY] = CLOSES[STALE_PROXY][:-2] + [None, None]

# public reference rows (reported_public, individually cited)
PUBLIC_ROWS = [
    ("PUB-001", "Total fund net TWR, FY ended Jun 30 2025", "LACERA Pension Plan", 0.097, "%",
     "ACFR 2025, Investment Results, printed p.110", "LACERA ACFR 2025"),
    ("PUB-002", "Total fund policy benchmark net TWR, FY ended Jun 30 2025", "LACERA Pension Plan",
     0.097, "%", "ACFR 2025, Investment Results, printed p.110", "LACERA ACFR 2025"),
    ("PUB-003", "Total investments (IBOR), Jun 30 2025", "LACERA Pension Plan", 85184786, "$K",
     "ACFR 2025, Investment Summary, printed p.108", "LACERA ACFR 2025"),
    ("PUB-004", "Total investments (IBOR), Jun 30 2025", "LACERA OPEB Master Trust", 5025968, "$K",
     "ACFR 2025, Investment Summary, printed p.109", "LACERA ACFR 2025"),
    ("PUB-005", "Annual money-weighted return net of investment expense, FY2025", "LACERA Pension Plan",
     0.097, "%", "ACFR 2025, RSI, printed p.95", "LACERA ACFR 2025"),
    ("PUB-006", "Total fund net return, quarter ended Mar 31 2026 (QTD)", "LACERA Pension Plan",
     0.006, "%", "Total Fund Performance Report Q1 2026, Summary, PDF p.6", "LACERA / Meketa"),
    ("PUB-007", "Total fund ending market value, Mar 31 2026", "LACERA Pension Plan", 89631, "$mm",
     "Total Fund Performance Report Q1 2026, PDF p.4", "LACERA / Meketa"),
    ("PUB-008", "Total fund AUM, May 2026", "LACERA Pension Plan", 93.9, "$B",
     "CIO Monthly Report July 2026, PDF p.8", "LACERA"),
]

# ---------------------------------------------------------------- python recomputation (QA baseline)
def recompute():
    bmv = {c: [0.0] * 12 for c in CAT_IDS}
    emv = {c: [0.0] * 12 for c in CAT_IDS}
    for c in CAT_IDS:
        bmv[c][0] = BMV0[c]
    for i in range(12):
        for c in CAT_IDS:
            emv[c][i] = bmv[c][i] * (1 + PORT_R[c][i]) + TRANSFERS[i][c]
            if i + 1 < 12:
                bmv[c][i + 1] = emv[c][i]
    bmv_tot = [sum(bmv[c][i] for c in CAT_IDS) for i in range(12)]
    emv_tot = [sum(emv[c][i] for c in CAT_IDS) for i in range(12)]
    w = {c: [bmv[c][i] / bmv_tot[i] for i in range(12)] for c in CAT_IDS}
    tf_r = [sum(w[c][i] * PORT_R[c][i] for c in CAT_IDS) for i in range(12)]
    idx = []
    g = 1.0
    for r in tf_r:
        g *= 1 + r
        idx.append(g)
    # benchmark
    def pol_w(i):
        month_start = MONTHS[i].replace(day=1)
        pol = [p for p in POLICY if p["effective"] <= month_start][-1]
        return pol
    bench_tf = [sum(pol_w(i)[c] * BENCH_R[c][i] for c in BENCH_R) for i in range(12)]
    bidx = []
    g = 1.0
    for r in bench_tf:
        g *= 1 + r
        bidx.append(g)
    qtd = idx[11] / idx[8] - 1
    fytd = idx[11] - 1
    m1 = tf_r[11]
    bqtd = bidx[11] / bidx[8] - 1
    bfytd = bidx[11] - 1
    bm1 = bench_tf[11]
    hurdle_m = (1 + HURDLE_ANNUAL) ** (1 / 12) - 1
    hurdle_q = (1 + HURDLE_ANNUAL) ** (3 / 12) - 1
    # contribution over Apr-Jun 2026 (i = 9..11)
    contrib = {c: sum(w[c][i] * PORT_R[c][i] for i in (9, 10, 11)) for c in CAT_IDS}
    arith = sum(contrib.values())
    residual = qtd - arith
    alloc = {c: emv[c][11] / emv_tot[11] for c in CAT_IDS}
    targets_now = [p for p in POLICY if p["effective"] <= AS_OF][-1]
    return {
        "bmv_total_first": bmv_tot[0], "emv_total_last": emv_tot[11],
        "tf_monthly": tf_r, "bench_monthly": bench_tf,
        "tf_1m": m1, "tf_qtd": qtd, "tf_fytd": fytd,
        "bench_1m": bm1, "bench_qtd": bqtd, "bench_fytd": bfytd,
        "hurdle_m": hurdle_m, "hurdle_q": hurdle_q, "hurdle_fy": HURDLE_ANNUAL,
        "contrib_qtd": contrib, "contrib_arith_total": arith, "contrib_residual": residual,
        "alloc_actual": alloc,
        "alloc_target": {c: targets_now.get(c, 0.0) for c in CAT_IDS},
        "weights_jun": {c: w[c][11] for c in CAT_IDS},
    }

EXPECTED = recompute()
with open(QA_JSON, "w", encoding="utf-8") as f:
    json.dump(EXPECTED, f, indent=1, default=str)

# ---------------------------------------------------------------- workbook helpers
wb = Workbook()

def style_hdr(ws, row, cols, start=1):
    for j in range(len(cols)):
        cell = ws.cell(row=row, column=start + j, value=cols[j])
        cell.font = F_HDR
        cell.fill = FILL_HDR
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")

def put(ws, row, col, value, font=F_FORMULA, fmt=None, fill=None, border=True, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if border:
        cell.border = BORDER
    if wrap:
        cell.alignment = WRAP
    return cell

def sheet_title(ws, title, note=None):
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    ws["A2"] = DISCLAIMER
    ws["A2"].font = F_NOTE
    ws["A2"].fill = FILL_DISC
    if note:
        ws["A3"] = note
        ws["A3"].font = F_NOTE

def col_widths(ws, widths, start=1):
    for j, wd in enumerate(widths):
        ws.column_dimensions[get_column_letter(start + j)].width = wd

# ================================================================ README
ws = wb.active
ws.title = "README"
sheet_title(ws, "Portfolio Analytics Dashboard Workbook — Prototype")
rows = [
    ("", ""),
    ("Purpose", "Auditable Excel bridge between the starter Market Pulse / ACFR toolkit and a "
                "future web dashboard prototype. Demonstrates classified inputs, valid return / "
                "contribution / allocation calculations, data-quality controls, the corrected "
                "ACFR trackers, and a normalized export contract."),
    ("Entity", "DEMOFUND — a wholly synthetic demonstration fund. No DEMOFUND value is, or is "
               "derived from, an actual LACERA figure."),
    ("As-of date", "2026-06-30 (named cell AsOfDate on Policy_Targets). Demo fiscal year: "
                   "July 2025 – June 2026."),
    ("Color legend", "BLUE font = hardcoded input. BLACK = same-sheet formula. GREEN = link to "
                     "another sheet. Amber banner = disclaimer. Status colors: green PASS, amber "
                     "WARN, red FAIL."),
    ("Classification legend", "reported_public = value quoted from a cited public source for the "
                              "exact stated period. synthetic = invented demo data. proxy_estimate "
                              "= public proxy for an exposure. calculated = derived by documented "
                              "formula (inherits weakest input class). stale / missing = data-state "
                              "overlays, tracked separately."),
    ("Sheet map", "Lists > Inputs_Public > Inputs_Market > Inputs_Portfolio > Policy_Targets > "
                  "Calc_Returns > Calc_Contribution > Calc_Allocation > Checks > Crosswalk > "
                  "QA_Checklist > Exec_View > Export_Contract"),
    ("Update workflow", "1) Edit BLUE input cells only. 2) Review Checks — every check must be "
                        "PASS or an explained WARN. 3) Review Exec_View. 4) Re-export "
                        "Export_Contract as CSV for the web prototype. 5) Record changes in "
                        "CHANGELOG.md."),
    ("Methodology", "See docs/workbook-methodology.md. Total return = sum of beginning-weight x "
                    "category return, chain-linked monthly (never an average of returns). "
                    "Contribution reconciles to the chain-linked total within the tolerance on "
                    "Calc_Contribution."),
    ("Known demo states", "DEMO-OIL is missing its final June close (missing-state demo). "
                          "DEMO-USD stops updating two business days early (stale-state demo). "
                          "Both are deliberate and flagged WARN on Checks."),
]
r = 4
for k, v in rows:
    if k:
        put(ws, r, 1, k, F_H2, border=False)
        put(ws, r, 2, v, F_FORMULA, border=False, wrap=True)
    r += 1
col_widths(ws, [24, 120])
for i in range(4, r):
    ws.row_dimensions[i].height = 30

# ================================================================ Lists
ws = wb.create_sheet("Lists")
sheet_title(ws, "Lists — controlled vocabularies")
lists = {
    "Status": ["Not Started", "In Progress", "Ready for Review", "Complete", "Blocked"],
    "Severity": ["High", "Medium", "Low"],
    "Frequency": ["Annual", "Quarterly", "Monthly", "Daily", "Ad Hoc"],
    "Classification": ["reported_public", "synthetic", "proxy_estimate", "calculated"],
    "DataState": ["current", "stale", "missing"],
    "Category": [c[1] for c in CATS],
    "PeriodType": ["D", "M", "Q", "FY", "1Y", "ITD"],
    "CheckStatus": ["PASS", "WARN", "FAIL"],
}
c0 = 1
for name, vals in lists.items():
    put(ws, 4, c0, name, F_HDR, fill=FILL_HDR)
    for i, v in enumerate(vals):
        put(ws, 5 + i, c0, v, F_INPUT, fill=FILL_INPUT)
    c0 += 1
col_widths(ws, [16] * len(lists))

# ================================================================ Inputs_Public
ws = wb.create_sheet("Inputs_Public")
sheet_title(ws, "Inputs_Public — cited public reference values (reported_public)",
            "Each row is quoted from a public LACERA document for the exact stated period. "
            "These values are context anchors only; no DEMOFUND calculation uses them.")
hdr = ["record_id", "metric", "entity", "value", "unit", "as_of_date", "period_start",
       "period_end", "period_type", "book_of_record", "return_method", "gross_net",
       "source (page/table)", "provider", "retrieved_date", "classification"]
style_hdr(ws, 5, hdr)
pub_meta = {
    "PUB-001": (dt.date(2025, 6, 30), dt.date(2024, 7, 1), dt.date(2025, 6, 30), "FY", "IBOR", "TWR", "net"),
    "PUB-002": (dt.date(2025, 6, 30), dt.date(2024, 7, 1), dt.date(2025, 6, 30), "FY", "n/a", "TWR", "net"),
    "PUB-003": (dt.date(2025, 6, 30), None, None, "FY", "IBOR", "n/a", "n/a"),
    "PUB-004": (dt.date(2025, 6, 30), None, None, "FY", "IBOR", "n/a", "n/a"),
    "PUB-005": (dt.date(2025, 6, 30), dt.date(2024, 7, 1), dt.date(2025, 6, 30), "FY", "n/a", "MWR", "net"),
    "PUB-006": (dt.date(2026, 3, 31), dt.date(2026, 1, 1), dt.date(2026, 3, 31), "Q", "n/a", "TWR", "net"),
    "PUB-007": (dt.date(2026, 3, 31), None, None, "Q", "IBOR", "n/a", "n/a"),
    "PUB-008": (dt.date(2026, 5, 31), None, None, "M", "IBOR", "n/a", "n/a"),
}
r = 6
for rid, metric, entity, value, unit, src, prov in PUBLIC_ROWS:
    asof, ps, pe, pt, book, rm, gn = pub_meta[rid]
    vals = [rid, metric, entity, value, unit, asof, ps, pe, pt, book, rm, gn, src, prov,
            dt.date(2026, 8, 4), "reported_public"]
    for j, v in enumerate(vals):
        fmt = None
        if j == 3:
            fmt = PCT2 if unit == "%" else MM0
        if j in (5, 6, 7, 14):
            fmt = DATE
        put(ws, r, 1 + j, v, F_INPUT, fmt=fmt, fill=FILL_INPUT, wrap=(j in (1, 12)))
    r += 1
for _rr in range(6, 6 + len(PUBLIC_ROWS)):
    ws.row_dimensions[_rr].height = 30
ws.freeze_panes = "A6"
col_widths(ws, [10, 44, 24, 12, 6, 11, 11, 11, 7, 8, 8, 7, 46, 16, 12, 15])

# ================================================================ Inputs_Market
ws = wb.create_sheet("Inputs_Market")
sheet_title(ws, "Inputs_Market — synthetic daily market-proxy closes (market context only)",
            "Synthetic series. NOT market data; never a portfolio return. Daily return column is "
            "calculated. DEMO-OIL missing final close and DEMO-USD staleness are deliberate.")
hdr = ["date", "proxy_id", "proxy_name", "category_read_through", "close", "daily_return",
       "classification", "source", "as_of_note"]
style_hdr(ws, 5, hdr)
r = 6
first_data_row = r
prev_row_for = {}
for di, day in enumerate(JUN_DAYS):
    for pid, name, cat, _p0 in PROXIES:
        close = CLOSES[pid][di]
        put(ws, r, 1, day, F_INPUT, fmt=DATE, fill=FILL_INPUT)
        put(ws, r, 2, pid, F_INPUT, fill=FILL_INPUT)
        put(ws, r, 3, name, F_INPUT, fill=FILL_INPUT)
        put(ws, r, 4, cat, F_INPUT, fill=FILL_INPUT)
        if close is None:
            put(ws, r, 5, "", F_INPUT, fill=FILL_INPUT)
        else:
            put(ws, r, 5, close, F_INPUT, fmt=PX4, fill=FILL_INPUT)
        prev = prev_row_for.get(pid)
        if prev is None:
            put(ws, r, 6, "", F_FORMULA)
        else:
            f = (f'=IF(OR(E{r}="",E{prev}=""),"",E{r}/E{prev}-1)')
            put(ws, r, 6, f, F_FORMULA, fmt=PCT2)
        put(ws, r, 7, "synthetic", F_INPUT, fill=FILL_INPUT)
        put(ws, r, 8, "synthetic generator seed 20260630", F_INPUT, fill=FILL_INPUT)
        put(ws, r, 9, "", F_INPUT, fill=FILL_INPUT)
        prev_row_for[pid] = r
        r += 1
last_market_row = r - 1
ws.freeze_panes = "A6"
col_widths(ws, [11, 16, 28, 26, 10, 11, 13, 30, 12])

# ================================================================ Inputs_Portfolio
ws = wb.create_sheet("Inputs_Portfolio")
sheet_title(ws, "Inputs_Portfolio — synthetic DEMOFUND monthly inputs",
            "All values synthetic. Returns are monthly net TWR-style category returns. "
            "Transfers are end-of-month internal reallocations (net zero across categories).")
put(ws, 4, 1, "Initial beginning market values ($mm), 1 Jul 2025", F_H2, border=False)
style_hdr(ws, 5, ["category_id", "category", "initial_bmv_mm", "classification"])
r = 6
BMV0_ROW = {}
for cid, cname in CATS:
    put(ws, r, 1, cid, F_INPUT, fill=FILL_INPUT)
    put(ws, r, 2, cname, F_INPUT, fill=FILL_INPUT)
    put(ws, r, 3, BMV0[cid], F_INPUT, fmt=MM1, fill=FILL_INPUT)
    put(ws, r, 4, "synthetic", F_INPUT, fill=FILL_INPUT)
    BMV0_ROW[cid] = r
    r += 1

base = r + 2
put(ws, base - 1, 1, "Monthly category net returns (decimal) — portfolio", F_H2, border=False)
style_hdr(ws, base, ["month_end"] + CAT_IDS)
PORT_R_ROW0 = base + 1
for i, mend in enumerate(MONTHS):
    put(ws, base + 1 + i, 1, mend, F_INPUT, fmt=MON, fill=FILL_INPUT)
    for j, cid in enumerate(CAT_IDS):
        put(ws, base + 1 + i, 2 + j, PORT_R[cid][i], F_INPUT, fmt=PCT2, fill=FILL_INPUT)

base2 = base + 15
put(ws, base2 - 1, 1, "Monthly category benchmark net returns (decimal) — synthetic benchmark", F_H2, border=False)
style_hdr(ws, base2, ["month_end", "GROWTH", "CREDIT", "RAIH", "RRM"])
BENCH_R_ROW0 = base2 + 1
for i, mend in enumerate(MONTHS):
    put(ws, base2 + 1 + i, 1, mend, F_INPUT, fmt=MON, fill=FILL_INPUT)
    for j, cid in enumerate(("GROWTH", "CREDIT", "RAIH", "RRM")):
        put(ws, base2 + 1 + i, 2 + j, BENCH_R[cid][i], F_INPUT, fmt=PCT2, fill=FILL_INPUT)

base3 = base2 + 15
put(ws, base3 - 1, 1, "End-of-month internal transfers ($mm, must net to zero per month)", F_H2, border=False)
style_hdr(ws, base3, ["month_end"] + CAT_IDS + ["net (check)"])
TRANS_ROW0 = base3 + 1
for i, mend in enumerate(MONTHS):
    rr = base3 + 1 + i
    put(ws, rr, 1, mend, F_INPUT, fmt=MON, fill=FILL_INPUT)
    for j, cid in enumerate(CAT_IDS):
        put(ws, rr, 2 + j, TRANSFERS[i][cid], F_INPUT, fmt=MM1, fill=FILL_INPUT)
    put(ws, rr, 2 + len(CAT_IDS), f"=SUM(B{rr}:G{rr})", F_FORMULA, fmt=MM1)
ws.freeze_panes = "A6"
col_widths(ws, [12, 30, 13, 13, 13, 13, 13, 13])
ws.cell(row=6, column=3).comment = Comment(
    "Synthetic starting scale (~$10bn total) chosen for readability; not related to any actual fund value.", "prototype")

# ================================================================ Policy_Targets
ws = wb.create_sheet("Policy_Targets")
sheet_title(ws, "Policy_Targets — effective-dated policy weights, ranges, hurdle",
            "Two policy versions demonstrate effective-dating. Weights apply to the four "
            "benchmark categories; Overlays/Other carry 0 policy weight.")
put(ws, 4, 1, "AsOfDate", F_H2, border=False)
put(ws, 4, 2, AS_OF, F_INPUT, fmt=DATE, fill=FILL_INPUT)
wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
    "AsOfDate", attr_text="Policy_Targets!$B$4"))
put(ws, 4, 4, "SchemaVersion", F_H2, border=False)
put(ws, 4, 5, "1.0.0", F_INPUT, fill=FILL_INPUT)
wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
    "SchemaVersion", attr_text="Policy_Targets!$E$4"))
put(ws, 5, 1, "Hurdle (annual, synthetic assumption)", F_H2, border=False)
put(ws, 5, 2, HURDLE_ANNUAL, F_INPUT, fmt=PCT2, fill=FILL_INPUT)
ws.cell(row=5, column=2).comment = Comment(
    "Synthetic actuarial-style hurdle for the demo; the actual assumed rates are published in the ACFR.", "prototype")

style_hdr(ws, 7, ["category_id", "target v1 (eff 2025-07-01)", "range +/- v1",
                  "target v2 (eff 2026-01-01)", "range +/- v2"])
put(ws, 8, 1, "effective_date", F_NOTE, border=False)
put(ws, 8, 2, POLICY[0]["effective"], F_INPUT, fmt=DATE, fill=FILL_INPUT)
put(ws, 8, 4, POLICY[1]["effective"], F_INPUT, fmt=DATE, fill=FILL_INPUT)
TGT_ROW0 = 9
for k, cid in enumerate(("GROWTH", "CREDIT", "RAIH", "RRM")):
    rr = TGT_ROW0 + k
    put(ws, rr, 1, cid, F_INPUT, fill=FILL_INPUT)
    put(ws, rr, 2, POLICY[0][cid], F_INPUT, fmt=PCT1, fill=FILL_INPUT)
    put(ws, rr, 3, RANGES[cid], F_INPUT, fmt=PCT1, fill=FILL_INPUT)
    put(ws, rr, 4, POLICY[1][cid], F_INPUT, fmt=PCT1, fill=FILL_INPUT)
    put(ws, rr, 5, RANGES[cid], F_INPUT, fmt=PCT1, fill=FILL_INPUT)
put(ws, TGT_ROW0 + 4, 1, "sum", F_NOTE, border=False)
put(ws, TGT_ROW0 + 4, 2, f"=SUM(B{TGT_ROW0}:B{TGT_ROW0+3})", F_FORMULA, fmt=PCT1)
put(ws, TGT_ROW0 + 4, 4, f"=SUM(D{TGT_ROW0}:D{TGT_ROW0+3})", F_FORMULA, fmt=PCT1)

put(ws, 15, 1, "Applied monthly policy weights (formula: version effective on month start)", F_H2, border=False)
style_hdr(ws, 16, ["month_end", "GROWTH", "CREDIT", "RAIH", "RRM"])
APPL_ROW0 = 17
for i, mend in enumerate(MONTHS):
    rr = APPL_ROW0 + i
    put(ws, rr, 1, mend, F_LINK, fmt=MON)
    ws.cell(row=rr, column=1).value = f"=Inputs_Portfolio!A{PORT_R_ROW0 + i}"
    for j in range(4):
        tr = TGT_ROW0 + j
        col = get_column_letter(2 + j)
        # month start = EOMONTH(month_end, -1)+1 ; v2 applies if month start >= v2 effective
        f = (f'=IF(EOMONTH($A{rr},-1)+1>=$D$8,'
             f'INDEX($D${TGT_ROW0}:$D${TGT_ROW0+3},{j+1}),'
             f'INDEX($B${TGT_ROW0}:$B${TGT_ROW0+3},{j+1}))')
        put(ws, rr, 2 + j, f, F_FORMULA, fmt=PCT1)
col_widths(ws, [26, 22, 12, 22, 12])

# ================================================================ Calc_Returns
ws = wb.create_sheet("Calc_Returns")
sheet_title(ws, "Calc_Returns — DEMOFUND monthly engine (all formulas)",
            "BMV evolves from inputs; weights are beginning-of-month; total return is the "
            "weighted sum of category returns, chain-linked through a growth index.")
NCAT = len(CAT_IDS)
def grid(ws, top, label, formula_fn, fmt, total_formula_fn=None):
    put(ws, top, 1, label, F_H2, border=False)
    style_hdr(ws, top + 1, ["month_end"] + CAT_IDS + (["TOTAL"] if total_formula_fn else []))
    for i in range(12):
        rr = top + 2 + i
        put(ws, rr, 1, f"=Inputs_Portfolio!A{PORT_R_ROW0 + i}", F_LINK, fmt=MON)
        for j in range(NCAT):
            put(ws, rr, 2 + j, formula_fn(i, j, rr), F_FORMULA, fmt=fmt)
        if total_formula_fn:
            put(ws, rr, 2 + NCAT, total_formula_fn(i, rr), F_FORMULA, fmt=fmt)
    return top + 2  # first data row

BMV_T = 5
EMV_T = BMV_T + 16
EMV_R0 = EMV_T + 2
bmv_r0 = grid(
    ws, BMV_T, "Beginning market value ($mm)",
    lambda i, j, rr: (f"=Inputs_Portfolio!C{BMV0_ROW[CAT_IDS[j]]}" if i == 0
                      else f"={get_column_letter(2+j)}{EMV_R0 + (i-1)}"),
    MM1, lambda i, rr: f"=SUM(B{rr}:G{rr})")
for i in range(1):
    rr = bmv_r0 + i
    for j in range(NCAT):
        ws.cell(row=rr, column=2 + j).font = F_LINK
emv_r0 = grid(
    ws, EMV_T, "Ending market value ($mm) = BMV x (1+return) + transfer",
    lambda i, j, rr: (f"={get_column_letter(2+j)}{bmv_r0+i}*(1+Inputs_Portfolio!{get_column_letter(2+j)}{PORT_R_ROW0+i})"
                      f"+Inputs_Portfolio!{get_column_letter(2+j)}{TRANS_ROW0+i}"),
    MM1, lambda i, rr: f"=SUM(B{rr}:G{rr})")
W_T = EMV_T + 16
w_r0 = grid(
    ws, W_T, "Beginning-of-month weight = BMV / total BMV",
    lambda i, j, rr: f"={get_column_letter(2+j)}{bmv_r0+i}/$H{bmv_r0+i}",
    PCT2, lambda i, rr: f"=SUM(B{rr}:G{rr})")
R_T = W_T + 16
r_r0 = grid(
    ws, R_T, "Category monthly return (link)",
    lambda i, j, rr: f"=Inputs_Portfolio!{get_column_letter(2+j)}{PORT_R_ROW0+i}",
    PCT2)
TF_T = R_T + 16
put(ws, TF_T, 1, "Total fund monthly return, growth indexes, benchmark", F_H2, border=False)
style_hdr(ws, TF_T + 1, ["month_end", "TF return = SUMPRODUCT(w,r)", "TF growth index",
                         "Bench TF return", "Bench growth index", "Hurdle growth index"])
TF_R0 = TF_T + 2
for i in range(12):
    rr = TF_R0 + i
    put(ws, rr, 1, f"=Inputs_Portfolio!A{PORT_R_ROW0 + i}", F_LINK, fmt=MON)
    put(ws, rr, 2, f"=SUMPRODUCT(B{w_r0+i}:G{w_r0+i},B{r_r0+i}:G{r_r0+i})", F_FORMULA, fmt=PCT2)
    prev_idx = f"C{rr-1}" if i else "1"
    put(ws, rr, 3, f"={prev_idx}*(1+B{rr})", F_FORMULA, fmt="0.0000")
    put(ws, rr, 4, (f"=SUMPRODUCT(Policy_Targets!B{APPL_ROW0+i}:E{APPL_ROW0+i},"
                    f"Inputs_Portfolio!B{BENCH_R_ROW0+i}:E{BENCH_R_ROW0+i})"), F_FORMULA, fmt=PCT2)
    prev_b = f"E{rr-1}" if i else "1"
    put(ws, rr, 5, f"={prev_b}*(1+D{rr})", F_FORMULA, fmt="0.0000")
    prev_h = f"F{rr-1}" if i else "1"
    put(ws, rr, 6, f"={prev_h}*(1+Policy_Targets!$B$5)^(1/12)", F_FORMULA, fmt="0.0000")
    # note: monthly hurdle growth = prev * (1+annual)^(1/12)
PER_T = TF_R0 + 13
put(ws, PER_T, 1, "Period results (as of AsOfDate)", F_H2, border=False)
style_hdr(ws, PER_T + 1, ["period", "portfolio", "benchmark", "hurdle", "excess vs bench"])
PER_R0 = PER_T + 2
jun, mar = TF_R0 + 11, TF_R0 + 8
rows_per = [
    ("1M (Jun 2026)", f"=B{jun}", f"=D{jun}", f"=(1+Policy_Targets!$B$5)^(1/12)-1"),
    ("QTD (Apr-Jun 2026)", f"=C{jun}/C{mar}-1", f"=E{jun}/E{mar}-1", f"=(1+Policy_Targets!$B$5)^(3/12)-1"),
    ("FYTD = 1Y (Jul 2025-Jun 2026)", f"=C{jun}-1", f"=E{jun}-1", "=Policy_Targets!$B$5"),
]
for k, (lbl, fp, fb, fh) in enumerate(rows_per):
    rr = PER_R0 + k
    put(ws, rr, 1, lbl, F_FORMULA)
    put(ws, rr, 2, fp, F_FORMULA, fmt=PCT2)
    put(ws, rr, 3, fb, F_FORMULA, fmt=PCT2)
    put(ws, rr, 4, fh, F_FORMULA, fmt=PCT2)
    put(ws, rr, 5, f"=B{rr}-C{rr}", F_FORMULA, fmt=PCT2)
for nm, addr in (("TF_1M", f"$B${PER_R0}"), ("TF_QTD", f"$B${PER_R0+1}"), ("TF_FYTD", f"$B${PER_R0+2}"),
                 ("BM_1M", f"$C${PER_R0}"), ("BM_QTD", f"$C${PER_R0+1}"), ("BM_FYTD", f"$C${PER_R0+2}")):
    wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
        nm, attr_text=f"Calc_Returns!{addr}"))
ws.freeze_panes = "B7"
col_widths(ws, [26, 13, 13, 13, 13, 13, 13, 13])

# ================================================================ Calc_Contribution
ws = wb.create_sheet("Calc_Contribution")
sheet_title(ws, "Calc_Contribution — QTD contribution with reconciliation",
            "Contribution = beginning-of-month weight x monthly return, summed Apr-Jun 2026. "
            "The compounding residual vs the chain-linked QTD return is disclosed and tested "
            "against the tolerance below.")
put(ws, 4, 1, "Tolerance (abs residual)", F_H2, border=False)
put(ws, 4, 2, 0.0010, F_INPUT, fmt=PCT2, fill=FILL_INPUT)
ws.cell(row=4, column=2).comment = Comment(
    "10 bps: monthly-arithmetic contributions vs chain-linked quarterly total.", "prototype")
style_hdr(ws, 6, ["category", "Apr w", "Apr r", "Apr c", "May w", "May r", "May c",
                  "Jun w", "Jun r", "Jun c", "QTD contribution"])
CON_R0 = 7
mon_rows = [9, 10, 11]  # Apr, May, Jun indexes
for k, cid in enumerate(CAT_IDS):
    rr = CON_R0 + k
    put(ws, rr, 1, CATS[k][1], F_LINK)
    for mi, midx in enumerate(mon_rows):
        cw = get_column_letter(2 + mi * 3)
        cr = get_column_letter(3 + mi * 3)
        cc = get_column_letter(4 + mi * 3)
        put(ws, rr, 2 + mi * 3, f"=Calc_Returns!{get_column_letter(2+k)}{w_r0+midx}", F_LINK, fmt=PCT2)
        put(ws, rr, 3 + mi * 3, f"=Calc_Returns!{get_column_letter(2+k)}{r_r0+midx}", F_LINK, fmt=PCT2)
        put(ws, rr, 4 + mi * 3, f"={cw}{rr}*{cr}{rr}", F_FORMULA, fmt=PCT2)
    put(ws, rr, 11, f"=D{rr}+G{rr}+J{rr}", F_FORMULA, fmt=PCT2)
SUM_R = CON_R0 + len(CAT_IDS)
put(ws, SUM_R, 1, "Arithmetic total (sum of contributions)", F_FORMULA)
put(ws, SUM_R, 11, f"=SUM(K{CON_R0}:K{SUM_R-1})", F_FORMULA, fmt=PCT2)
put(ws, SUM_R + 1, 1, "Chain-linked QTD return (Calc_Returns)", F_LINK)
put(ws, SUM_R + 1, 11, "=TF_QTD", F_LINK, fmt=PCT2)
put(ws, SUM_R + 2, 1, "Compounding residual (chain-linked minus arithmetic)", F_FORMULA)
put(ws, SUM_R + 2, 11, f"=K{SUM_R+1}-K{SUM_R}", F_FORMULA, fmt=PCT2)
put(ws, SUM_R + 3, 1, "Reconciliation status", F_FORMULA)
put(ws, SUM_R + 3, 11, f'=IF(ABS(K{SUM_R+2})<=B4,"PASS","FAIL")', F_FORMULA)
ws.conditional_formatting.add(
    f"K{SUM_R+3}",
    CellIsRule(operator="equal", formula=['"PASS"'], fill=FILL_PASS))
ws.conditional_formatting.add(
    f"K{SUM_R+3}",
    CellIsRule(operator="equal", formula=['"FAIL"'], fill=FILL_FAIL))
wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
    "Contribution_Residual", attr_text=f"Calc_Contribution!$K${SUM_R+2}"))
wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
    "Contribution_Status", attr_text=f"Calc_Contribution!$K${SUM_R+3}"))
ws.freeze_panes = "B7"
col_widths(ws, [40] + [9] * 9 + [14])

# ================================================================ Calc_Allocation
ws = wb.create_sheet("Calc_Allocation")
sheet_title(ws, "Calc_Allocation — actual vs effective-dated policy targets (as of AsOfDate)")
style_hdr(ws, 5, ["category", "EMV $mm (Jun 2026)", "actual %", "target % (current policy)",
                  "over/under %", "over/under $mm", "range +/-", "range status"])
AL_R0 = 6
jun_emv_row = emv_r0 + 11
for k, cid in enumerate(CAT_IDS):
    rr = AL_R0 + k
    put(ws, rr, 1, CATS[k][1], F_LINK)
    put(ws, rr, 2, f"=Calc_Returns!{get_column_letter(2+k)}{jun_emv_row}", F_LINK, fmt=MM1)
    put(ws, rr, 3, f"=B{rr}/$B${AL_R0+len(CAT_IDS)}", F_FORMULA, fmt=PCT2)
    if cid in ("GROWTH", "CREDIT", "RAIH", "RRM"):
        j = ("GROWTH", "CREDIT", "RAIH", "RRM").index(cid)
        put(ws, rr, 4, (f'=IF(AsOfDate>=Policy_Targets!$D$8,Policy_Targets!D{TGT_ROW0+j},'
                        f'Policy_Targets!B{TGT_ROW0+j})'), F_LINK, fmt=PCT1)
        put(ws, rr, 7, f"=Policy_Targets!C{TGT_ROW0+j}", F_LINK, fmt=PCT1)
        put(ws, rr, 8, f'=IF(ABS(E{rr})<=G{rr},"WITHIN RANGE","OUT OF RANGE")', F_FORMULA)
    else:
        put(ws, rr, 4, 0, F_INPUT, fmt=PCT1, fill=FILL_INPUT)
        put(ws, rr, 7, "", F_FORMULA)
        put(ws, rr, 8, "n/a (no policy weight)", F_FORMULA)
    put(ws, rr, 5, f"=C{rr}-D{rr}", F_FORMULA, fmt=PCT2)
    put(ws, rr, 6, f"=E{rr}*$B${AL_R0+len(CAT_IDS)}", F_FORMULA, fmt=MM1)
TOT_R = AL_R0 + len(CAT_IDS)
put(ws, TOT_R, 1, "Total", F_FORMULA)
put(ws, TOT_R, 2, f"=SUM(B{AL_R0}:B{TOT_R-1})", F_FORMULA, fmt=MM1)
put(ws, TOT_R, 3, f"=SUM(C{AL_R0}:C{TOT_R-1})", F_FORMULA, fmt=PCT2)
put(ws, TOT_R, 4, f"=SUM(D{AL_R0}:D{TOT_R-1})", F_FORMULA, fmt=PCT1)
ws.conditional_formatting.add(
    f"H{AL_R0}:H{TOT_R-1}",
    CellIsRule(operator="equal", formula=['"OUT OF RANGE"'], fill=FILL_FAIL))
ws.freeze_panes = "A6"
col_widths(ws, [32, 16, 10, 20, 12, 14, 10, 20])

# ================================================================ Checks
ws = wb.create_sheet("Checks")
sheet_title(ws, "Checks — data-quality and reconciliation controls",
            "Every check must be PASS, or WARN with an explanation. The two WARNs below are "
            "deliberate demo states (missing / stale market data).")
style_hdr(ws, 5, ["check_id", "description", "result", "status", "expected", "note"])
CHK_R0 = 6
w_last = w_r0 + 11
checks = [
    ("CHK-01", "Beginning weights sum to 100% every month (max abs deviation)",
     f"=MAX(ABS(Calc_Returns!H{w_r0}-1)," + ",".join(
         f"ABS(Calc_Returns!H{w_r0+i}-1)" for i in range(1, 12)) + ")",
     "=IF(C{r}<0.000001,\"PASS\",\"FAIL\")", "< 0.0001%", ""),
    ("CHK-02", "Contribution residual within tolerance",
     "=ABS(Contribution_Residual)",
     "=IF(C{r}<=Calc_Contribution!B4,\"PASS\",\"FAIL\")", "<= 10 bps", ""),
    ("CHK-03", "Allocation actual % sums to 100%",
     f"=ABS(Calc_Allocation!C{TOT_R}-1)",
     "=IF(C{r}<0.000001,\"PASS\",\"FAIL\")", "< 0.0001%", ""),
    ("CHK-04", "Policy target weights sum to 100% (both versions)",
     f"=MAX(ABS(Policy_Targets!B{TGT_ROW0+4}-1),ABS(Policy_Targets!D{TGT_ROW0+4}-1))",
     "=IF(C{r}<0.000001,\"PASS\",\"FAIL\")", "< 0.0001%", ""),
    ("CHK-05", "Internal transfers net to zero every month",
     f"=MAX(" + ",".join(f"ABS(Inputs_Portfolio!H{TRANS_ROW0+i})" for i in range(12)) + ")",
     "=IF(C{r}<0.005,\"PASS\",\"FAIL\")", "< 0.005 $mm", ""),
    ("CHK-06", "Market strip completeness (missing closes)",
     f"={len(JUN_DAYS)*len(PROXIES)}-COUNT(Inputs_Market!E{first_data_row}:E{last_market_row})",
     "=IF(C{r}=0,\"PASS\",\"WARN\")", "0 missing",
     "3 missing by design: DEMO-OIL final close; DEMO-USD last two closes (stale demo)"),
    ("CHK-07", "Market strip freshness (proxies with data on final business day)",
     f"=COUNTIFS(Inputs_Market!A{first_data_row}:A{last_market_row},MAX(Inputs_Market!A{first_data_row}:A{last_market_row}),"
     f"Inputs_Market!E{first_data_row}:E{last_market_row},\"<>\")",
     f"=IF(C{{r}}={len(PROXIES)},\"PASS\",\"WARN\")", f"{len(PROXIES)} proxies",
     "DEMO-OIL and DEMO-USD lack final-day closes by design"),
    ("CHK-08", "Benchmark return coverage = 12 months x 4 categories",
     f"=COUNT(Inputs_Portfolio!B{BENCH_R_ROW0}:E{BENCH_R_ROW0+11})",
     "=IF(C{r}=48,\"PASS\",\"FAIL\")", "48", ""),
    ("CHK-09", "Public reference rows fully sourced (blank source/as-of cells)",
     f"=COUNTBLANK(Inputs_Public!M6:M{5+len(PUBLIC_ROWS)})+COUNTBLANK(Inputs_Public!F6:F{5+len(PUBLIC_ROWS)})",
     "=IF(C{r}=0,\"PASS\",\"FAIL\")", "0", ""),
    ("CHK-10", "Crosswalk item count (structured reference)",
     "=ROWS(ACFRCrosswalkTable[Section])",
     "=IF(C{r}=23,\"PASS\",\"FAIL\")", "23", "Fixes starter Overview!B10 header miscount"),
    ("CHK-11", "QA control count (structured reference)",
     "=ROWS(ACFRQATable[Category])",
     "=IF(C{r}=28,\"PASS\",\"FAIL\")", "28", ""),
    ("CHK-12", "Export record count matches expected",
     '=COUNTIF(Export_Contract!A:A,"REC-*")',
     "=IF(C{r}=EXPORT_EXPECTED,\"PASS\",\"FAIL\")", "see Export sheet", ""),
]
for k, (cid_, desc, result_f, status_f, exp, note) in enumerate(checks):
    rr = CHK_R0 + k
    put(ws, rr, 1, cid_, F_FORMULA)
    put(ws, rr, 2, desc, F_FORMULA, wrap=True)
    put(ws, rr, 3, result_f, F_FORMULA, fmt="0.00000")
    put(ws, rr, 4, status_f.replace("{r}", str(rr)), F_FORMULA)
    put(ws, rr, 5, exp, F_FORMULA)
    put(ws, rr, 6, note, F_NOTE, wrap=True)
for state, fill in (("PASS", FILL_PASS), ("WARN", FILL_WARNF), ("FAIL", FILL_FAIL)):
    ws.conditional_formatting.add(
        f"D{CHK_R0}:D{CHK_R0+len(checks)-1}",
        CellIsRule(operator="equal", formula=[f'"{state}"'], fill=fill))
put(ws, CHK_R0 + len(checks) + 1, 2, "PASS count", F_H2, border=False)
put(ws, CHK_R0 + len(checks) + 1, 3,
    f'=COUNTIF(D{CHK_R0}:D{CHK_R0+len(checks)-1},"PASS")', F_FORMULA)
put(ws, CHK_R0 + len(checks) + 2, 2, "WARN count", F_H2, border=False)
put(ws, CHK_R0 + len(checks) + 2, 3,
    f'=COUNTIF(D{CHK_R0}:D{CHK_R0+len(checks)-1},"WARN")', F_FORMULA)
put(ws, CHK_R0 + len(checks) + 3, 2, "FAIL count", F_H2, border=False)
put(ws, CHK_R0 + len(checks) + 3, 3,
    f'=COUNTIF(D{CHK_R0}:D{CHK_R0+len(checks)-1},"FAIL")', F_FORMULA)
CHECK_COUNTS_ROW = CHK_R0 + len(checks) + 1
for _cr in range(CHECK_COUNTS_ROW, CHECK_COUNTS_ROW + 3):
    ws.row_dimensions[_cr].height = 16
wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
    "Checks_Pass", attr_text=f"Checks!$C${CHECK_COUNTS_ROW}"))
wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
    "Checks_Warn", attr_text=f"Checks!$C${CHECK_COUNTS_ROW+1}"))
wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
    "Checks_Fail", attr_text=f"Checks!$C${CHECK_COUNTS_ROW+2}"))
ws.freeze_panes = "A6"
col_widths(ws, [9, 52, 12, 9, 12, 52])

# ================================================================ Crosswalk (corrected)
CROSSWALK_ROWS = [
    (104, "Investment", "Chief Investment Officer's Report", "Pension Plan & OPEB",
     "AUM, fiscal-year return, market context and narrative",
     "Final performance package; approved AUM", "Investment Summary and Financial Statements",
     "June 30 fiscal year-end",
     "All headline amounts and returns agree to the final tables; narrative uses approved figures."),
    (108, "Investment", "Investment Summary — Pension Plan", "Pension Plan",
     "Fair value and percent of total by functional category and sub-asset class",
     "Investment Book of Record (IBOR)", "Statement of Fiduciary Net Position / ABOR",
     "As of June 30",
     "Total investments agrees to IBOR; percentages equal 100.0%; explain IBOR/ABOR difference."),
    (109, "Investment", "Investment Summary — OPEB Master Trust", "OPEB Master Trust",
     "Fair value and percent of total by functional category and sub-asset class",
     "Investment Book of Record (IBOR)", "Statement of Fiduciary Net Position / ABOR",
     "As of June 30",
     "Total investments agrees to IBOR; operational cash treatment verified; percentages equal 100.0%."),
    (109, "Investment", "Investment Summary — OPEB Custodial Fund", "OPEB Custodial Fund",
     "Cash and fixed-income fair value and allocation",
     "Investment Book of Record (IBOR)", "Statement of Fiduciary Net Position / ABOR",
     "As of June 30",
     "Cash plus fixed income equals total; percentages equal 100.0%; fund not mixed with Master Trust."),
    (110, "Investment", "Investment Results Based on Fair Value — Pension Plan", "Pension Plan",
     "Quarter, 1-, 3-, 5- and 10-year net returns and policy benchmarks",
     "Final performance system / custodian performance book", "Quarterly Total Fund Performance Report",
     "Periods ended June 30",
     "Net-of-fees labels, annualization, benchmark mapping and delayed-asset footnotes verified."),
    (111, "Investment", "Investment Results Based on Fair Value — OPEB Master Trust", "OPEB Master Trust",
     "Quarter, 1-, 3-, 5- and 10-year net returns and policy benchmarks",
     "Final performance system / custodian performance book", "Quarterly OPEB Performance Report",
     "Periods ended June 30",
     "Net-of-fees labels, annualization, benchmark mapping, new-account treatment and lags verified."),
    (112, "Investment", "Total Investment Rates of Return — Pension Plan", "Pension Plan",
     "10-year fair value, TWRR, MWRR, smoothed return, assumed return and funded ratio",
     "Final performance package; actuarial valuation", "Investment Summary; Actuarial Section",
     "Fiscal years ended June 30",
     "Historical values roll forward correctly; current-year N/A fields and actuarial timing are documented."),
    (113, "Investment", "Total Investment Rates of Return — OPEB Master Trust", "OPEB Master Trust",
     "Historical fair value, TWRR, MWRR, smoothed return, assumed return and funded ratio",
     "Final performance package; actuarial valuation", "Investment Summary; Actuarial Section",
     "Fiscal years ended June 30",
     "Historical values roll forward; schedule length and unavailable current-year actuarial metrics are documented."),
    (114, "Investment", "Largest Equity Holdings — Pension Plan", "Pension Plan",
     "Top equity holdings by fair value and shares",
     "Custodian holdings file", "Published Investment Holdings", "As of June 30",
     "Top holdings sorted by fair value; shares, issuer names and custody-only scope verified."),
    (114, "Investment", "Largest Equity Holdings — OPEB Master Trust", "OPEB Master Trust",
     "Top equity holdings by fair value and shares",
     "Custodian holdings file", "Published Investment Holdings", "As of June 30",
     "Top holdings sorted by fair value; shares, issuer names and custody-only scope verified."),
    (115, "Investment", "Largest Fixed Income Holdings — Pension Plan", "Pension Plan",
     "Top fixed-income holdings by fair value and par",
     "Custodian holdings file", "Published Investment Holdings", "As of June 30",
     "Top holdings sorted by fair value; par, coupon, maturity and security descriptions verified."),
    (115, "Investment", "Largest Fixed Income Holdings — OPEB Master Trust", "OPEB Master Trust",
     "Top fixed-income holdings by fair value and par",
     "Custodian holdings file", "Published Investment Holdings", "As of June 30",
     "Top holdings sorted by fair value; par, coupon, maturity and security descriptions verified."),
    (116, "Investment", "Schedule of Investment Management Fees", "Pension Plan & OPEB",
     "Management fees by mandate/category",
     "Accounts payable / fee system / general ledger", "Investment cost report and manager contracts",
     "Fiscal year",
     "Fees reconcile to accounting records; classification, accruals and fund allocation verified."),
    (117, "Investment", "List of Investment Managers", "Pension Plan & OPEB",
     "Manager roster by asset category",
     "Manager master / contract database", "Fee schedule and quarterly scorecards",
     "As of publication cutoff",
     "All active managers included; terminated managers handled consistently; names and categories standardized."),
    (95, "Required Supplementary Information", "Schedule of Investment Returns — Pension Plan", "Pension Plan",
     "Annual money-weighted return net of investment expense",
     "Final performance package", "Investment Section rates-of-return table", "Fiscal year",
     "Return methodology and net-of-expense definition agree across sections."),
    (99, "Required Supplementary Information", "Schedule of Investment Returns — OPEB Trust", "OPEB Trust",
     "Annual money-weighted return net of investment expense",
     "Final performance package", "Investment Section rates-of-return table", "Fiscal year",
     "Return methodology and net-of-expense definition agree across sections."),
    (101, "Supplementary Information", "Schedule of Investment Expenses", "Pension Plan & OPEB",
     "Investment expenses by type",
     "General ledger / accounts payable", "Investment management fee schedule", "Fiscal year",
     "Expense totals reconcile to audited financial statements and fee disclosures."),
    (101, "Supplementary Information", "Schedule of Payments to Consultants — Pension Plan", "Pension Plan",
     "Consultant payments",
     "General ledger / accounts payable", "Contract and vendor master", "Fiscal year",
     "Vendor names, amounts, service classifications and fiscal-year cutoff verified."),
    (25, "Financial", "Statement of Fiduciary Net Position — Investments", "Pension Plan & OPEB",
     "Investments at fair value and fiduciary net position",
     "Accounting Book of Record (ABOR)", "IBOR-to-ABOR reconciliation", "As of June 30",
     "ABOR totals agree to audited financial statements; reconciling items to IBOR are documented."),
    (27, "Financial", "Statement of Changes in Fiduciary Net Position — Investment Income and Expenses",
     "Pension Plan & OPEB", "Net appreciation, interest, dividends and investment expense",
     "General ledger / ABOR", "Performance and investment-expense schedules", "Fiscal year",
     "Investment income and expenses reconcile to audited statements and supporting schedules."),
    (60, "Financial Notes", "Note G — Deposit and Investment Risks", "Pension Plan & OPEB",
     "Custodial credit, concentration, interest-rate, currency and other investment risks",
     "Custodian risk reports / investment records", "Investment Policy Statement and holdings",
     "As of June 30",
     "Risk disclosures agree to holdings and policy; classifications and material concentrations reviewed."),
    (155, "Statistical", "Changes in Fiduciary Net Position — Pension Plan", "Pension Plan",
     "Ten-year additions, deductions and net-position history",
     "Audited financial statements", "Financial Section", "Ten fiscal years",
     "Current year agrees to financial statements; prior years unchanged; units and signs consistent."),
    (156, "Statistical", "Changes in Fiduciary Net Position — OPEB Trust", "OPEB Trust",
     "Historical additions, deductions and net-position data",
     "Audited financial statements", "Financial Section", "Available fiscal years",
     "Current year agrees to financial statements; prior years unchanged; Master/Custodial scopes understood."),
]
DEMO_STATUS = (["Complete"] * 4 + ["Ready for Review"] * 2 + ["In Progress"] * 4 +
               ["Not Started"] * 12 + ["Blocked"])
ws = wb.create_sheet("Crosswalk")
sheet_title(ws, "ACFR Investment Data Crosswalk — corrected tracker",
            "Structure follows the ACFR 2025 table of contents (printed page numbers). Statuses "
            "and due dates are ILLUSTRATIVE demo values. Owners intentionally TBD.")
hdr = ["ACFR page", "Section", "Table / disclosure", "Fund", "Primary metric(s)",
       "Authoritative source", "Secondary tie-out", "Data owner", "Reviewer",
       "Reporting period", "Due date", "Status", "Validation / tie-out test",
       "Variance / issue", "Notes", "Source URL"]
style_hdr(ws, 5, hdr)
XW_R0 = 6
due0 = dt.date(2026, 9, 15)
for k, row in enumerate(CROSSWALK_ROWS):
    page, section, table, fund, metrics, src, tie, period, test = row
    rr = XW_R0 + k
    vals = [page, section, table, fund, metrics, src, tie, "TBD — confirm internally", "TBD",
            period, due0 + dt.timedelta(days=7 * (k % 8)), DEMO_STATUS[k], test, "", "",
            "https://www.lacera.gov/sites/default/files/assets/documents/annual_reports/ACFR-2025.pdf"]
    for j, v in enumerate(vals):
        fmt = "0" if j == 0 else (DATE if j == 10 else None)
        put(ws, rr, 1 + j, v, F_INPUT, fmt=fmt, fill=FILL_INPUT, wrap=(j in (2, 4, 5, 6, 12)))
XW_LAST = XW_R0 + len(CROSSWALK_ROWS) - 1
tab = Table(displayName="ACFRCrosswalkTable", ref=f"A5:P{XW_LAST}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)
dv = DataValidation(type="list", formula1="Lists!$A$5:$A$9", allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"L{XW_R0}:L{XW_LAST}")
for state, fill in (("Complete", FILL_PASS), ("Blocked", FILL_FAIL)):
    ws.conditional_formatting.add(
        f"L{XW_R0}:L{XW_LAST}",
        CellIsRule(operator="equal", formula=[f'"{state}"'], fill=fill))
ws.conditional_formatting.add(
    f"L{XW_R0}:L{XW_LAST}",
    FormulaRule(formula=[f'OR($L{XW_R0}="In Progress",$L{XW_R0}="Ready for Review")'], fill=FILL_WARNF))
for _rr in range(XW_R0, XW_LAST + 1):
    ws.row_dimensions[_rr].height = 42
ws.freeze_panes = "C6"
col_widths(ws, [9, 22, 40, 18, 34, 30, 28, 20, 10, 18, 11, 15, 44, 14, 14, 40])

# ================================================================ QA_Checklist
QA_ROWS = [
    ("Scope & dates", "Confirm every table uses the correct June 30 reporting date.", "All schedules", "Annual", "High"),
    ("Scope & dates", "Keep Pension Plan, OPEB Master Trust and OPEB Custodial Fund data separate.", "All schedules", "Annual", "High"),
    ("Scope & dates", "Confirm dollars-in-thousands labels and scaling are consistent.", "Financial and investment tables", "Annual", "High"),
    ("Market values", "Tie Investment Summary totals to the final IBOR extract.", "Investment Summary", "Annual", "High"),
    ("Market values", "Prepare and retain the IBOR-to-ABOR reconciliation instead of forcing the two books to match.", "Investment Summary / Financial Statements", "Annual", "High"),
    ("Market values", "Verify allocation percentages calculate to 100.0%, subject to stated exclusions.", "Investment Summary", "Annual", "High"),
    ("Market values", "Validate signs for currency hedges, overlay positions and other assets.", "Investment Summary", "Annual", "High"),
    ("Market values", "Confirm operational cash treatment in the OPEB Master Trust schedule.", "OPEB Investment Summary", "Annual", "High"),
    ("Performance", "Confirm all published returns are net of manager fees where labeled.", "Investment Results", "Annual", "High"),
    ("Performance", "Recalculate excess return versus each policy benchmark.", "Investment Results", "Annual", "High"),
    ("Performance", "Verify quarter, one-, three-, five- and ten-year periods end on the same date.", "Investment Results", "Annual", "High"),
    ("Performance", "Check annualization methodology for periods longer than one year.", "Investment Results", "Annual", "High"),
    ("Performance", "Verify benchmark names, custom benchmark definitions and effective dates.", "Investment Results", "Annual", "High"),
    ("Performance", "Document one- or three-month reporting lags for applicable private assets and benchmarks.", "Investment Results", "Annual", "Medium"),
    ("Performance", "Confirm blanks, dashes and N/A fields are used consistently for unavailable history.", "Return schedules", "Annual", "Medium"),
    ("Historical schedules", "Roll prior-year columns forward without altering published historical values.", "Ten-year schedules", "Annual", "High"),
    ("Historical schedules", "Tie current-year fair value and returns to the current Investment Summary and performance package.", "Ten-year schedules", "Annual", "High"),
    ("Holdings", "Sort top holdings by fair value and confirm the requested count.", "Largest holdings", "Annual", "Medium"),
    ("Holdings", "Validate shares/par, issuer descriptions, coupon and maturity fields.", "Largest holdings", "Annual", "High"),
    ("Holdings", "Confirm the holdings schedules reflect assets held in custody and disclose that scope.", "Largest holdings", "Annual", "Medium"),
    ("Fees & expenses", "Reconcile management fees to accounts payable, accruals and the investment cost report.", "Fee schedules", "Annual", "High"),
    ("Fees & expenses", "Check that fees and expenses are allocated to the correct fund and category.", "Fee schedules", "Annual", "High"),
    ("Managers", "Reconcile the manager list to active contracts and quarterly scorecards.", "Manager list", "Annual", "Medium"),
    ("Cross-report", "Tie ACFR performance figures to the final June quarterly performance package.", "Investment Section", "Annual", "High"),
    ("Cross-report", "Tie investment income and expenses to the audited financial statements.", "Financial / Investment", "Annual", "High"),
    ("Cross-report", "Confirm PAFR investment figures are a faithful summary of the final ACFR.", "ACFR / PAFR", "Annual", "Medium"),
    ("Governance", "Archive the final source files, query parameters and review evidence.", "All schedules", "Annual", "High"),
    ("Governance", "Obtain preparer and independent reviewer sign-off before publication.", "All schedules", "Annual", "High"),
]
QA_STATUS = (["Complete"] * 6 + ["In Progress"] * 3 + ["Ready for Review"] * 2 +
             ["Not Started"] * 16 + ["Blocked"])
ws = wb.create_sheet("QA_Checklist")
sheet_title(ws, "ACFR Investment QA Checklist — corrected tracker",
            "Statuses are ILLUSTRATIVE demo values.")
hdr = ["Category", "Control / check", "Applies to", "Frequency", "Severity", "Owner", "Status",
       "Evidence / file", "Resolution notes"]
style_hdr(ws, 5, hdr)
QA_R0 = 6
for k, (cat, ctrl, appl, freq, sev) in enumerate(QA_ROWS):
    rr = QA_R0 + k
    vals = [cat, ctrl, appl, freq, sev, "TBD", QA_STATUS[k], "", ""]
    for j, v in enumerate(vals):
        put(ws, rr, 1 + j, v, F_INPUT, fill=FILL_INPUT, wrap=(j == 1))
QA_LAST = QA_R0 + len(QA_ROWS) - 1
tab = Table(displayName="ACFRQATable", ref=f"A5:I{QA_LAST}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)
for col_letter, src_range in (("D", "Lists!$C$5:$C$9"), ("E", "Lists!$B$5:$B$7"), ("G", "Lists!$A$5:$A$9")):
    dv = DataValidation(type="list", formula1=src_range, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{QA_R0}:{col_letter}{QA_LAST}")
for state, fill in (("Complete", FILL_PASS), ("Blocked", FILL_FAIL)):
    ws.conditional_formatting.add(
        f"G{QA_R0}:G{QA_LAST}",
        CellIsRule(operator="equal", formula=[f'"{state}"'], fill=fill))
ws.conditional_formatting.add(
    f"G{QA_R0}:G{QA_LAST}",
    FormulaRule(formula=[f'OR($G{QA_R0}="In Progress",$G{QA_R0}="Ready for Review")'], fill=FILL_WARNF))
for _rr in range(QA_R0, QA_LAST + 1):
    ws.row_dimensions[_rr].height = 28
ws.freeze_panes = "C6"
col_widths(ws, [20, 60, 26, 11, 10, 10, 15, 18, 18])

# ================================================================ Exec_View
ws = wb.create_sheet("Exec_View")
sheet_title(ws, "Executive View — DEMOFUND (synthetic) as of 2026-06-30")
put(ws, 4, 1, "All portfolio figures on this sheet are SYNTHETIC (classification: synthetic / "
              "calculated). Net-of-fees TWR-style monthly linking. See Calc sheets for lineage.",
    F_WARN, border=False, wrap=True)
ws.merge_cells("A4:H4")
put(ws, 6, 1, "Illustrative performance (net)", F_H2, border=False)
style_hdr(ws, 7, ["period", "DEMOFUND", "policy benchmark (synthetic)", "hurdle (synthetic)",
                  "excess vs benchmark"])
for k in range(3):
    rr = 8 + k
    src = PER_R0 + k
    put(ws, rr, 1, f"=Calc_Returns!A{src}", F_LINK)
    put(ws, rr, 2, f"=Calc_Returns!B{src}", F_LINK, fmt=PCT2)
    put(ws, rr, 3, f"=Calc_Returns!C{src}", F_LINK, fmt=PCT2)
    put(ws, rr, 4, f"=Calc_Returns!D{src}", F_LINK, fmt=PCT2)
    put(ws, rr, 5, f"=Calc_Returns!E{src}", F_LINK, fmt=PCT2)
put(ws, 12, 1, "QTD contribution by category (reconciles on Calc_Contribution)", F_H2, border=False)
style_hdr(ws, 13, ["category", "QTD contribution", "", "reconciliation", "value"])
for k in range(len(CAT_IDS)):
    rr = 14 + k
    put(ws, rr, 1, f"=Calc_Contribution!A{CON_R0+k}", F_LINK)
    put(ws, rr, 2, f"=Calc_Contribution!K{CON_R0+k}", F_LINK, fmt=PCT2)
put(ws, 14, 4, "Arithmetic total", F_LINK)
put(ws, 14, 5, f"=Calc_Contribution!K{SUM_R}", F_LINK, fmt=PCT2)
put(ws, 15, 4, "Chain-linked QTD", F_LINK)
put(ws, 15, 5, "=TF_QTD", F_LINK, fmt=PCT2)
put(ws, 16, 4, "Compounding residual", F_LINK)
put(ws, 16, 5, "=Contribution_Residual", F_LINK, fmt=PCT2)
put(ws, 17, 4, "Status", F_LINK)
put(ws, 17, 5, "=Contribution_Status", F_LINK)
put(ws, 22, 1, "Allocation vs policy (as of 2026-06-30)", F_H2, border=False)
style_hdr(ws, 23, ["category", "actual %", "target %", "over/under %", "range status"])
for k in range(len(CAT_IDS)):
    rr = 24 + k
    put(ws, rr, 1, f"=Calc_Allocation!A{AL_R0+k}", F_LINK)
    put(ws, rr, 2, f"=Calc_Allocation!C{AL_R0+k}", F_LINK, fmt=PCT2)
    put(ws, rr, 3, f"=Calc_Allocation!D{AL_R0+k}", F_LINK, fmt=PCT1)
    put(ws, rr, 4, f"=Calc_Allocation!E{AL_R0+k}", F_LINK, fmt=PCT2)
    put(ws, rr, 5, f"=Calc_Allocation!H{AL_R0+k}", F_LINK)
put(ws, 31, 1, "Data trust", F_H2, border=False)
put(ws, 32, 1, "Checks PASS / WARN / FAIL", F_LINK)
put(ws, 32, 2, '=Checks_Pass&" / "&Checks_Warn&" / "&Checks_Fail', F_LINK)
put(ws, 33, 1, "Contribution reconciliation", F_LINK)
put(ws, 33, 2, "=Contribution_Status", F_LINK)
put(ws, 34, 1, "Known demo data states", F_FORMULA)
put(ws, 34, 2, "DEMO-OIL missing final close; DEMO-USD stale (deliberate)", F_FORMULA, wrap=True)
put(ws, 36, 1, "ACFR readiness (illustrative statuses)", F_H2, border=False)
put(ws, 37, 1, "Crosswalk items / complete", F_FORMULA)
put(ws, 37, 2, '=ROWS(ACFRCrosswalkTable[Section])&" / "&COUNTIF(ACFRCrosswalkTable[Status],"Complete")', F_FORMULA)
put(ws, 38, 1, "QA controls / complete", F_FORMULA)
put(ws, 38, 2, '=ROWS(ACFRQATable[Category])&" / "&COUNTIF(ACFRQATable[Status],"Complete")', F_FORMULA)
put(ws, 39, 1, "Crosswalk completion %", F_FORMULA)
put(ws, 39, 2, '=COUNTIF(ACFRCrosswalkTable[Status],"Complete")/ROWS(ACFRCrosswalkTable[Section])', F_FORMULA, fmt=PCT1)
put(ws, 41, 1, "Market context (synthetic proxies, final business day) — NOT portfolio performance", F_H2, border=False)
style_hdr(ws, 42, ["proxy", "category read-through", "last daily return", "state"])
for k, (pid, name, cat, _p0) in enumerate(PROXIES):
    rr = 43 + k
    # last row for proxy pid in Inputs_Market: rows are interleaved; compute last day row
    last_r = first_data_row + (len(JUN_DAYS) - 1) * len(PROXIES) + k
    put(ws, rr, 1, name, F_FORMULA)
    put(ws, rr, 2, cat, F_FORMULA)
    put(ws, rr, 3, f"=Inputs_Market!F{last_r}", F_LINK, fmt=PCT2)
    put(ws, rr, 4, f'=IF(Inputs_Market!E{last_r}="","missing/stale","current")', F_FORMULA)
ws.conditional_formatting.add(
    f"D43:D{42+len(PROXIES)}",
    CellIsRule(operator="equal", formula=['"missing/stale"'], fill=FILL_WARNF))
col_widths(ws, [34, 24, 26, 22, 20, 12, 12, 12])

# ================================================================ Export_Contract
ws = wb.create_sheet("Export_Contract")
sheet_title(ws, "Export_Contract — normalized interface for the web prototype (schema 1.0.0)",
            "One record per row. Derived values are formula-linked to their calculation cells; "
            "input values are static. This sheet, exported as CSV, is the only interface the web "
            "prototype consumes.")
hdr = ["record_id", "record_type", "entity_id", "metric_id", "category_id", "value", "unit",
       "currency", "scale", "as_of_date", "period_start", "period_end", "period_type",
       "frequency", "classification", "source_type", "source_name", "page_table", "provider",
       "retrieved_date", "book_of_record", "return_method", "gross_net", "valuation_status",
       "benchmark_id", "method_id", "quality_status", "note", "schema_version"]
EXPORT_HEADER_ROWS = 6  # title rows + header row (rows 1..6 in col A used before data)
style_hdr(ws, 6, hdr)
r = 7
def ex_row(record_type, metric_id, category_id, value, unit, asof, ps, pe, pt, freq,
           classification, source_type, source_name, page_table, provider,
           book="n/a", rm="n/a", gn="n/a", vs="final", bench="", method="", qs="ok", note="",
           value_is_formula=False, fmt=PCT2, entity="DEMOFUND"):
    global r
    rid = f"REC-{r-6:04d}"
    scale = {"$mm": "mm", "$K": "k", "$B": "bn"}.get(unit, "1")
    vals = [rid, record_type, entity, metric_id, category_id, value, unit, "USD",
            scale, asof, ps, pe, pt, freq, classification,
            source_type, source_name, page_table, provider, dt.date(2026, 8, 4),
            book, rm, gn, vs, bench, method, qs, note, "1.0.0"]
    for j, v in enumerate(vals):
        fnt = F_LINK if (j == 5 and value_is_formula) else (F_FORMULA if value_is_formula else F_INPUT)
        fmt_j = None
        if j == 5:
            fmt_j = fmt
        if j in (9, 10, 11, 19):
            fmt_j = DATE
        put(ws, r, 1 + j, v, fnt, fmt=fmt_j,
            fill=None if value_is_formula else FILL_INPUT, border=False)
    r += 1

# monthly portfolio category + total returns (calculated -> formula links)
for i, mend in enumerate(MONTHS):
    ps = mend.replace(day=1)
    for j, cid in enumerate(CAT_IDS):
        ex_row("monthly_return", "net_return_m", cid,
               f"=Calc_Returns!{get_column_letter(2+j)}{r_r0+i}", "%", mend, ps, mend, "M", "Monthly",
               "synthetic", "workbook", "Inputs_Portfolio", "monthly returns grid", "synthetic generator",
               book="IBOR", rm="TWR", gn="net", method="monthly_net_return", value_is_formula=True)
    ex_row("monthly_return", "net_return_m", "TOTAL",
           f"=Calc_Returns!B{TF_R0+i}", "%", mend, ps, mend, "M", "Monthly",
           "calculated", "workbook", "Calc_Returns", "TF monthly SUMPRODUCT", "workbook formulas",
           book="IBOR", rm="TWR", gn="net", method="bop_weighted_sum", value_is_formula=True)
    for j, cid in enumerate(("GROWTH", "CREDIT", "RAIH", "RRM")):
        ex_row("monthly_benchmark_return", "bench_return_m", cid,
               f"=Inputs_Portfolio!{get_column_letter(2+j)}{BENCH_R_ROW0+i}", "%", mend, ps, mend, "M",
               "Monthly", "synthetic", "workbook", "Inputs_Portfolio", "benchmark grid",
               "synthetic generator", rm="TWR", gn="net", bench=f"BM-{cid}", value_is_formula=True)
    ex_row("monthly_benchmark_return", "bench_return_m", "TOTAL",
           f"=Calc_Returns!D{TF_R0+i}", "%", mend, ps, mend, "M", "Monthly",
           "calculated", "workbook", "Calc_Returns", "policy-weighted benchmark", "workbook formulas",
           rm="TWR", gn="net", bench="BM-TOTAL", method="policy_weighted_sum", value_is_formula=True)

# period aggregates
period_rows = [
    ("1M", dt.date(2026, 6, 1), AS_OF, f"=Calc_Returns!B{PER_R0}", f"=Calc_Returns!C{PER_R0}", f"=Calc_Returns!D{PER_R0}"),
    ("QTD", dt.date(2026, 4, 1), AS_OF, f"=Calc_Returns!B{PER_R0+1}", f"=Calc_Returns!C{PER_R0+1}", f"=Calc_Returns!D{PER_R0+1}"),
    ("FYTD", dt.date(2025, 7, 1), AS_OF, f"=Calc_Returns!B{PER_R0+2}", f"=Calc_Returns!C{PER_R0+2}", f"=Calc_Returns!D{PER_R0+2}"),
]
for pt, ps, pe, fp, fb, fh in period_rows:
    ex_row("period_return", "net_return", "TOTAL", fp, "%", AS_OF, ps, pe, pt, "Monthly",
           "calculated", "workbook", "Calc_Returns", "period results", "workbook formulas",
           book="IBOR", rm="TWR", gn="net", method="chain_linked", value_is_formula=True)
    ex_row("period_return", "bench_return", "TOTAL", fb, "%", AS_OF, ps, pe, pt, "Monthly",
           "calculated", "workbook", "Calc_Returns", "period results", "workbook formulas",
           rm="TWR", gn="net", bench="BM-TOTAL", method="chain_linked", value_is_formula=True)
    ex_row("period_return", "hurdle_return", "TOTAL", fh, "%", AS_OF, ps, pe, pt, "Monthly",
           "synthetic", "workbook", "Policy_Targets", "hurdle assumption", "synthetic assumption",
           method="geometric_scaling", value_is_formula=True)

# allocation records
for k, cid in enumerate(CAT_IDS):
    rr_src = AL_R0 + k
    ex_row("allocation", "emv", cid, f"=Calc_Allocation!B{rr_src}", "$mm", AS_OF, None, None, "M",
           "Monthly", "calculated", "workbook", "Calc_Allocation", "EMV link", "workbook formulas",
           book="IBOR", value_is_formula=True, fmt=MM1)
    ex_row("allocation", "weight_actual", cid, f"=Calc_Allocation!C{rr_src}", "%", AS_OF, None, None,
           "M", "Monthly", "calculated", "workbook", "Calc_Allocation", "actual %", "workbook formulas",
           value_is_formula=True)
    ex_row("allocation", "weight_target", cid, f"=Calc_Allocation!D{rr_src}", "%", AS_OF, None, None,
           "M", "Monthly", "synthetic", "workbook", "Policy_Targets", "effective-dated target",
           "synthetic assumption", value_is_formula=True)
    ex_row("allocation", "over_under_pct", cid, f"=Calc_Allocation!E{rr_src}", "%", AS_OF, None, None,
           "M", "Monthly", "calculated", "workbook", "Calc_Allocation", "over/under", "workbook formulas",
           value_is_formula=True)

# contribution records
for k, cid in enumerate(CAT_IDS):
    ex_row("contribution_qtd", "contribution", cid, f"=Calc_Contribution!K{CON_R0+k}", "%", AS_OF,
           dt.date(2026, 4, 1), AS_OF, "QTD", "Quarterly", "calculated", "workbook",
           "Calc_Contribution", "QTD contribution", "workbook formulas",
           method="bop_weight_x_return_sum", value_is_formula=True)
ex_row("contribution_qtd", "contribution_arith_total", "TOTAL", f"=Calc_Contribution!K{SUM_R}", "%",
       AS_OF, dt.date(2026, 4, 1), AS_OF, "QTD", "Quarterly", "calculated", "workbook",
       "Calc_Contribution", "arithmetic total", "workbook formulas", value_is_formula=True)
ex_row("contribution_qtd", "return_chain_linked", "TOTAL", "=TF_QTD", "%", AS_OF,
       dt.date(2026, 4, 1), AS_OF, "QTD", "Quarterly", "calculated", "workbook",
       "Calc_Returns", "chain-linked QTD", "workbook formulas", method="chain_linked",
       value_is_formula=True)
ex_row("contribution_qtd", "residual", "TOTAL", "=Contribution_Residual", "%", AS_OF,
       dt.date(2026, 4, 1), AS_OF, "QTD", "Quarterly", "calculated", "workbook",
       "Calc_Contribution", "compounding residual", "workbook formulas", value_is_formula=True)

# market close records (inputs -> static values)
for di, day in enumerate(JUN_DAYS):
    for pid, _name, cat, _p0 in PROXIES:
        close = CLOSES[pid][di]
        qs = "ok" if close is not None else "missing"
        ex_row("market_close", "close", pid, close if close is not None else "", "px", day, None,
               None, "D", "Daily", "synthetic", "workbook", "Inputs_Market",
               "synthetic proxy series", "synthetic generator", qs=qs,
               note=cat, fmt=PX4)

# public reference records (static). entity_id names the cited public entity, not DEMOFUND:
# these are quotations about a real fund and must never be attributed to the synthetic entity.
for (rid, metric, entity, value, unit, src, prov), _m in zip(PUBLIC_ROWS, range(len(PUBLIC_ROWS))):
    asof, ps, pe, pt, book, rm, gn = pub_meta[rid]
    ex_row("public_reference", metric, "", value, unit, asof, ps, pe, pt, "Ad Hoc",
           "reported_public", "public_report", src.split(",")[0], src, prov, book=book, rm=rm,
           gn=gn, note=f"quoted from {src}",
           fmt=PCT2 if unit == "%" else MM0, entity=entity)

# check result records (formula links to status)
for k, (cid_, desc, _rf, _sf, _exp, _note) in enumerate(checks):
    ex_row("check_result", cid_, "TOTAL", f"=Checks!D{CHK_R0+k}", "status", AS_OF, None, None, "M",
           "Monthly", "calculated", "workbook", "Checks", desc[:60], "workbook formulas",
           value_is_formula=True, fmt=None)

EXPORT_LAST = r - 1
N_RECORDS = EXPORT_LAST - 6
wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
    "EXPORT_EXPECTED", attr_text=str(N_RECORDS)))
ws.freeze_panes = "A7"
col_widths(ws, [10, 20, 11, 22, 16, 12, 6, 6, 5, 11, 11, 11, 7, 9, 15, 12, 18, 28, 20, 11,
                8, 8, 7, 9, 10, 22, 9, 26, 9])

wb.save(OUT)
print(f"saved {OUT}")
print(f"export records: {N_RECORDS}")
print("expected:", {k: (round(v, 6) if isinstance(v, float) else v)
                    for k, v in EXPECTED.items() if isinstance(v, (int, float))})
