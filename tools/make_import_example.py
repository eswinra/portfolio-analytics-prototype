"""Build a daily-workflow import example: outputs/Daily_Import_Example.xlsx + .csv.

Takes the canonical Pension fixture, appends one new market day (2026-07-01) in which every
proxy is priced (so importing it clears the missing/stale demo issues and lifts coverage from
40.5% to 43.5%), flips the two market controls to PASS to match, and packages the result as
(a) a ready-to-upload contract CSV and (b) an Excel workbook that walks the daily workflow.
"""
import csv
import io
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "data", "sample", "demofund_export_v1.csv")
OUT_CSV = os.path.join(ROOT, "outputs", "daily_import_example.csv")
OUT_XLSX = os.path.join(ROOT, "outputs", "Daily_Import_Example.xlsx")

NEW_DAY = "2026-07-01"
# plausible next-day moves per proxy (decimal)
MOVES = {
    "DEMO-EQ-GLOBAL": 0.0062,
    "DEMO-EQ-SMALL": 0.0085,
    "DEMO-BOND-AGG": -0.0012,
    "DEMO-GOLD": 0.0030,
    "DEMO-OIL": 0.0185,  # rebounds after its missing close
    "DEMO-USD": -0.0008,  # updates again after going stale
}

with io.open(SRC, encoding="utf-8") as f:
    reader = csv.reader(f)
    hdr = next(reader)
    rows = list(reader)

i = {name: hdr.index(name) for name in hdr}

# last present close per proxy
last_close: dict[str, float] = {}
for r in rows:
    if r[i["record_type"]] == "market_close" and r[i["value"]]:
        last_close[r[i["category_id"]]] = float(r[i["value"]])

# template a market row per proxy from its most recent row
template: dict[str, list[str]] = {}
for r in rows:
    if r[i["record_type"]] == "market_close":
        template[r[i["category_id"]]] = list(r)

max_rec = max(int(r[i["record_id"]].split("-")[1]) for r in rows)
new_rows: list[list[str]] = []
for n, (proxy, move) in enumerate(MOVES.items(), start=1):
    row = list(template[proxy])
    row[i["record_id"]] = f"REC-{max_rec + n:04d}"
    row[i["as_of_date"]] = NEW_DAY
    row[i["value"]] = format(round(last_close[proxy] * (1 + move), 4), ".10g")
    row[i["quality_status"]] = "ok"
    row[i["retrieved_date"]] = "2026-08-07"
    # schema 1.2 provenance: the day's paste is a user-import entry, not yet reviewed —
    # importing therefore demos V19 (entered_by) plus the draft banner and review workflow
    row[i["source_type"]] = "user_import"
    row[i["entered_by"]] = "PA-ANALYST-2"
    row[i["reviewed_by"]] = ""
    row[i["review_status"]] = "draft"
    new_rows.append(row)

# the daily workflow re-runs the checks: with a complete fresh day, CHK-06/07 pass
for r in rows:
    if r[i["record_type"]] == "check_result" and r[i["metric_id"]] in ("CHK-06", "CHK-07"):
        r[i["value"]] = "PASS"

all_rows = rows + new_rows
with io.open(OUT_CSV, "w", encoding="utf-8", newline="") as f:
    w = csv.writer(f, lineterminator="\n")
    w.writerow(hdr)
    w.writerows(all_rows)
print(f"wrote {OUT_CSV} ({len(all_rows)} records)")

# ---------------------------------------------------------------- Excel workbook
F_TITLE = Font(bold=True, size=14, color="1F3864")
F_H2 = Font(bold=True, size=11, color="1F3864")
F_HDR = Font(bold=True, size=10, color="FFFFFF")
F_TXT = Font(size=10)
F_NOTE = Font(italic=True, size=9, color="595959")
F_INPUT = Font(color="0000CC", size=10)
FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_NEW = PatternFill("solid", fgColor="E2EFDA")
FILL_INPUT = PatternFill("solid", fgColor="DDEBF7")
FILL_DISC = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# README
ws = wb.active
ws.title = "README"
ws["A1"] = "Daily import example — Fund Pulse prototype"
ws["A1"].font = F_TITLE
ws["A2"] = ("SYNTHETIC DEMO DATA. This example shows the daily end-of-day workflow and produces "
            "a file the prototype's Import view accepts. Not an official LACERA system.")
ws["A2"].font = F_NOTE
ws["A2"].fill = FILL_DISC
steps = [
    ("What this shows", "One new market day (2026-07-01) arrives in which every proxy is priced "
     "— including DEMO-OIL (previously missing its close) and DEMO-USD (previously stale). "
     "Importing the file clears both data issues and lifts policy-weight coverage from 40.5% "
     "to 43.5%."),
    ("Daily workflow", "1) At EOD, refresh or paste the new closes into the Daily_Closes sheet "
     "(blue cells). In a real internal setup this area is a Bloomberg BDH refresh.  "
     "2) Append the six generated contract rows to Export_Contract (here already appended — "
     "green rows at the bottom).  3) Re-run your checks (CHK-06/07 return to PASS with a "
     "complete, fresh strip).  4) Save Export_Contract as CSV — or use the companion "
     "daily_import_example.csv, which is exactly that export.  5) In the app: Import → drop the "
     "CSV → review the preflight (364 rows, 0 errors) → Apply."),
    ("What to look for after Apply", "Overview: 'Today's proxy pulse — market data through "
     "2026-07-01', read-through led by Global Equity, coverage 43.5%, both market issues "
     "cleared. A yellow DRAFT DATA banner appears: the six new rows are review_status=draft "
     "(entered by PA-ANALYST-2, not yet reviewed) — Exceptions shows one informational issue "
     "and the Team Activity panel now lists both actors. That is the schema-1.2 review "
     "workflow: a reviewer would set the rows reviewed/published in the source file."),
    ("Real-data caution", "For licensed market data (Bloomberg terminal exports), use this "
     "workflow only on an internal copy of the app and never commit exports to the public "
     "repository. Files are parsed entirely in the browser and never transmitted."),
]
r = 4
for k, v in steps:
    ws.cell(row=r, column=1, value=k).font = F_H2
    c = ws.cell(row=r, column=2, value=v)
    c.font = F_TXT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 66
    r += 1
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 110

# Daily_Closes — the analyst-facing input area
ws = wb.create_sheet("Daily_Closes")
ws["A1"] = "New end-of-day closes — 2026-07-01 (blue = the analyst's daily input)"
ws["A1"].font = F_TITLE
headers = ["proxy_id", "prior close", "new close (EOD input)", "daily move"]
for j, h in enumerate(headers):
    c = ws.cell(row=3, column=1 + j, value=h)
    c.font = F_HDR
    c.fill = FILL_HDR
    c.border = BORDER
rr = 4
for proxy, move in MOVES.items():
    prior = last_close[proxy]
    new = round(prior * (1 + move), 4)
    ws.cell(row=rr, column=1, value=proxy).font = F_TXT
    c = ws.cell(row=rr, column=2, value=prior)
    c.number_format = "0.0000"
    c.font = F_TXT
    c = ws.cell(row=rr, column=3, value=new)
    c.number_format = "0.0000"
    c.font = F_INPUT
    c.fill = FILL_INPUT
    c = ws.cell(row=rr, column=4, value=f"=C{rr}/B{rr}-1")
    c.number_format = "0.00%"
    for j in range(1, 5):
        ws.cell(row=rr, column=j).border = BORDER
    rr += 1
ws.cell(row=rr + 1, column=1,
        value="Prior closes for DEMO-OIL / DEMO-USD are their last present values "
              "(2026-06-29 / 2026-06-26).").font = F_NOTE
for col, wdt in (("A", 18), ("B", 12), ("C", 20), ("D", 12)):
    ws.column_dimensions[col].width = wdt

# Export_Contract — full table, new rows highlighted
ws = wb.create_sheet("Export_Contract")
ws["A1"] = "Contract export — save this sheet as CSV to import (green rows = today's new data)"
ws["A1"].font = F_TITLE
for j, h in enumerate(hdr):
    c = ws.cell(row=3, column=1 + j, value=h)
    c.font = F_HDR
    c.fill = FILL_HDR
new_ids = {r_[i["record_id"]] for r_ in new_rows}
changed_checks = {"CHK-06", "CHK-07"}
for rn, row in enumerate(all_rows, start=4):
    is_new = row[i["record_id"]] in new_ids
    is_chk = row[i["record_type"]] == "check_result" and row[i["metric_id"]] in changed_checks
    for j, v in enumerate(row):
        c = ws.cell(row=rn, column=1 + j, value=v)
        c.font = F_TXT
        if is_new or is_chk:
            c.fill = FILL_NEW
ws.freeze_panes = "A4"
for j in range(len(hdr)):
    ws.column_dimensions[get_column_letter(j + 1)].width = 13

wb.save(OUT_XLSX)
print(f"wrote {OUT_XLSX}")
