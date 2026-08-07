"""Build outputs/Fund_Pulse_Dashboard.xlsx — a shareable Excel mirror of the web dashboard.

One presentation tab per fund (KPIs, growth-of-$1 line chart, monthly-return bars, daily
read-through bars, allocation vs IPS bands, contribution reconciliation, pulse strip,
trend windows), an IPS policy tab, and per-fund Data tabs that feed every chart — so the
file keeps working as a dashboard if rows are appended later. Everything derives from the
same fixtures the web app bundles. Synthetic data only; safe to share.
"""
import csv
import io
import os
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "outputs", "Fund_Pulse_Dashboard.xlsx")

FUNDS = {
    "PENSION": {
        "label": "Pension fund (DEMOFUND)",
        "csv": os.path.join(ROOT, "data", "sample", "demofund_export_v1.csv"),
        "bands": {"GROWTH": (0.40, 0.56), "CREDIT": (0.09, 0.17), "RAIH": (0.11, 0.19), "RRM": (0.16, 0.32)},
        "weights": {"DEMO-EQ-GLOBAL": 0.305, "DEMO-BOND-AGG": 0.10, "DEMO-OIL": 0.03},
        "ips": ("invest_policy_stmt.pdf", "Tables 1-2, printed pp. 20-21"),
    },
    "OPEB": {
        "label": "OPEB fund (DEMO-OPEB)",
        "csv": os.path.join(ROOT, "data", "sample", "demo_opeb_export_v1.csv"),
        "bands": {"GROWTH": (0.35, 0.55), "CREDIT": (0.11, 0.21), "RAIH": (0.09, 0.17), "RRM": (0.17, 0.35)},
        "weights": {"DEMO-EQ-GLOBAL": 0.40, "DEMO-BOND-AGG": 0.145, "DEMO-OIL": 0.02},
        "ips": ("IPS-OPEB.pdf", "Tables 1-2, printed pp. 21-22"),
    },
}
CAT_LABEL = {"GROWTH": "Growth", "CREDIT": "Credit", "RAIH": "Real Assets & Infl. Hedges",
             "RRM": "Risk Reduction & Mitigation", "OVERLAY": "Overlays & Hedges", "OTHER": "Other Asset"}
CATS = ["GROWTH", "CREDIT", "RAIH", "RRM", "OVERLAY", "OTHER"]

DISCLAIMER = ("PROTOTYPE - SYNTHETIC DATA ONLY. Not an official LACERA system, performance "
              "report, or statement of endorsement. Generated from the same fixtures as the web "
              "prototype; the custodian remains the official book of record.")

# ---------------------------------------------------------------- load + derive per fund
def load(fund):
    with io.open(FUNDS[fund]["csv"], encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    d = {}
    monthly = {}
    bench = {}
    for r in rows:
        t = r["record_type"]
        if t == "monthly_return" and r["category_id"] == "TOTAL":
            monthly[r["period_end"]] = float(r["value"])
        elif t == "monthly_benchmark_return" and r["category_id"] == "TOTAL":
            bench[r["period_end"]] = float(r["value"])
    months = sorted(monthly)
    d["months"] = months
    d["port"] = [monthly[m] for m in months]
    d["bench"] = [bench.get(m) for m in months]
    gi, g = [], 1.0
    for r_ in d["port"]:
        g *= 1 + r_
        gi.append(g)
    bi, g = [], 1.0
    for r_ in d["bench"]:
        g *= 1 + (r_ or 0)
        bi.append(g)
    d["pidx"], d["bidx"] = gi, bi

    per = {}
    for r in rows:
        if r["record_type"] == "period_return":
            per.setdefault(r["period_type"], {})[r["metric_id"]] = float(r["value"])
    d["periods"] = per

    contrib, extras = [], {}
    for r in rows:
        if r["record_type"] == "contribution_qtd":
            if r["metric_id"] == "contribution":
                contrib.append((r["category_id"], float(r["value"])))
            else:
                extras[r["metric_id"]] = float(r["value"])
    contrib.sort(key=lambda x: CATS.index(x[0]))
    d["contrib"], d["cextras"] = contrib, extras

    alloc = {}
    for r in rows:
        if r["record_type"] == "allocation":
            alloc.setdefault(r["category_id"], {})[r["metric_id"]] = float(r["value"]) if r["value"] else None
    d["alloc"] = alloc

    closes = {}
    for r in rows:
        if r["record_type"] == "market_close":
            closes.setdefault(r["category_id"], []).append(
                (r["as_of_date"], float(r["value"]) if r["value"] else None))
    for v in closes.values():
        v.sort()
    d["closes"] = closes

    checks = {r["metric_id"]: r["value"] for r in rows if r["record_type"] == "check_result"}
    d["checks"] = checks
    return d


def present(series):
    return [(dt, c) for dt, c in series if c is not None]


def trend(series):
    s = present(series)
    n = len(s)
    def chg(k):
        return s[-1][1] / s[-1 - k][1] - 1 if n >= k + 1 else None
    mtd = None
    if n >= 2:
        month = s[-1][0][:7]
        inm = [c for dt, c in s if dt[:7] == month]
        if len(inm) >= 2:
            mtd = s[-1][1] / inm[0] - 1
    vol = None
    if n >= 21:
        rets = [s[i][1] / s[i - 1][1] - 1 for i in range(n - 20, n)]
        mean = sum(rets) / len(rets)
        vol = (sum((r - mean) ** 2 for r in rets) / (len(rets) - 1)) ** 0.5
    return {"last": s[-1][0] if s else None, "obs": n, "d1": chg(1), "d5": chg(5), "mtd": mtd, "vol20": vol}


def read_through_days(closes, weights):
    by = {}
    for pid, w in weights.items():
        s = present(closes.get(pid, []))
        for i in range(1, len(s)):
            r_ = s[i][1] / s[i - 1][1] - 1
            e = by.setdefault(s[i][0], [0.0, 0.0])
            e[0] += w * r_
            e[1] += w
    return sorted((dt, imp, cov) for dt, (imp, cov) in by.items())


def issues(d):
    out = []
    all_dates = sorted({dt for v in d["closes"].values() for dt, _ in v})
    last_day = all_dates[-1] if all_dates else None
    for pid, series in sorted(d["closes"].items()):
        s = present(series)
        if not s:
            continue
        age = (date.fromisoformat(last_day) - date.fromisoformat(s[-1][0])).days if last_day else 0
        if age > 3:
            out.append(f"{pid} - stale since {s[-1][0]} (control CHK-07)")
        elif series[-1][1] is None:
            out.append(f"{pid} - final close missing (control CHK-06)")
    return out

# ---------------------------------------------------------------- styles
F_TITLE = Font(bold=True, size=15, color="1F3864")
F_H2 = Font(bold=True, size=11, color="1F3864")
F_KPI = Font(bold=True, size=18, color="14202B")
F_LBL = Font(size=8, color="5D6B76", bold=True)
F_HDR = Font(bold=True, size=9, color="FFFFFF")
F_TXT = Font(size=10)
F_NOTE = Font(italic=True, size=8.5, color="595959")
FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_TILE = PatternFill("solid", fgColor="F5F3EC")
FILL_DISC = PatternFill("solid", fgColor="FFF2CC")
FILL_PASS = PatternFill("solid", fgColor="C6EFCE")
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT2, PCT1 = "0.00%", "0.0%"

wb = Workbook()

def put(ws, r, c, v, font=F_TXT, fmt=None, fill=None, border=False, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if border:
        cell.border = BORDER
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell

# ---------------------------------------------------------------- README
ws = wb.active
ws.title = "README"
put(ws, 1, 1, "Fund Pulse — Excel dashboard (synthetic demo)", F_TITLE)
put(ws, 2, 1, DISCLAIMER, F_NOTE, fill=FILL_DISC)
notes = [
    ("What this is", "A shareable Excel mirror of the Fund Pulse web prototype: one dashboard "
     "tab per fund plus the IPS policy reference. All values derive from the same synthetic "
     "fixtures the web app bundles (data/sample/*.csv)."),
    ("Live version", "https://eswinra.github.io/portfolio-analytics-prototype/"),
    ("Future use as a dashboard", "Every chart reads from the Data_PENSION / Data_OPEB tabs. "
     "Append rows there (new months or new market days) and extend the chart ranges to keep "
     "using this file as a living dashboard. The web app's Import page accepts the same data "
     "as CSV."),
    ("Regenerate", "python tools/make_dashboard_workbook.py (repo: eswinra/"
     "portfolio-analytics-prototype)."),
    ("Method notes", "Total return = weighted sum of category returns, chain-linked monthly - "
     "never averaged. Daily read-through = policy half-step weight x proxy daily return over "
     "covered liquid proxies only; private classes are excluded (IPS benchmarks lagged 1-3 "
     "months). sigma-20 = std deviation of the last 20 daily returns, not annualized."),
]
r = 4
for k, v in notes:
    put(ws, r, 1, k, F_H2)
    put(ws, r, 2, v, F_TXT, wrap=True)
    ws.row_dimensions[r].height = 42
    r += 1
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 105

# ---------------------------------------------------------------- per-fund data + dashboard
for fund, cfg in FUNDS.items():
    d = load(fund)
    tr_days = read_through_days(d["closes"], cfg["weights"])
    iss = issues(d)

    # ---- Data sheet
    dws = wb.create_sheet(f"Data_{fund}")
    put(dws, 1, 1, f"Data — {cfg['label']} (charts read these ranges; append rows to extend)", F_H2)
    hdrs = ["month_end", "portfolio_return", "benchmark_return", "portfolio_index",
            "benchmark_index", "ret_pos", "ret_neg"]
    for j, h in enumerate(hdrs):
        put(dws, 3, 1 + j, h, F_HDR, fill=FILL_HDR)
    for k, m in enumerate(d["months"]):
        rr = 4 + k
        put(dws, rr, 1, m)
        put(dws, rr, 2, d["port"][k], fmt=PCT2)
        put(dws, rr, 3, d["bench"][k], fmt=PCT2)
        put(dws, rr, 4, d["pidx"][k], fmt="0.0000")
        put(dws, rr, 5, d["bidx"][k], fmt="0.0000")
        put(dws, rr, 6, d["port"][k] if d["port"][k] >= 0 else None, fmt=PCT2)
        put(dws, rr, 7, d["port"][k] if d["port"][k] < 0 else None, fmt=PCT2)
    n_m = len(d["months"])
    rt0 = 4 + n_m + 2
    put(dws, rt0 - 1, 1, "daily read-through (covered liquid proxies)", F_H2)
    for j, h in enumerate(["date", "impact", "coverage", "imp_pos", "imp_neg"]):
        put(dws, rt0, 1 + j, h, F_HDR, fill=FILL_HDR)
    for k, (dt, imp, cov) in enumerate(tr_days):
        rr = rt0 + 1 + k
        put(dws, rr, 1, dt)
        put(dws, rr, 2, imp, fmt="0.000%")
        put(dws, rr, 3, cov, fmt=PCT1)
        put(dws, rr, 4, imp if imp >= 0 else None, fmt="0.000%")
        put(dws, rr, 5, imp if imp < 0 else None, fmt="0.000%")
    for col, wd in (("A", 12), ("B", 12), ("C", 12), ("D", 12), ("E", 12), ("F", 10), ("G", 10)):
        dws.column_dimensions[col].width = wd
    dws.sheet_state = "visible"

    # ---- Dashboard sheet
    ws = wb.create_sheet(f"{fund}_Dashboard")
    put(ws, 1, 1, f"Fund Pulse — {cfg['label']}", F_TITLE)
    put(ws, 2, 1, DISCLAIMER, F_NOTE, fill=FILL_DISC)
    asof = d["months"][-1] if d["months"] else "n/a"
    put(ws, 3, 1, f"As of {asof} · synthetic proxy view · net of fees, TWR-style monthly linking", F_NOTE)

    # KPI band
    per = d["periods"]
    fytd, qtd = per.get("FYTD", {}), per.get("QTD", {})
    kpis = [
        ("FISCAL YTD RETURN (NET)", fytd.get("net_return"), PCT2,
         f"vs benchmark {fytd.get('bench_return', 0):.2%}"),
        ("EXCESS VS BENCHMARK (FYTD)", (fytd.get("net_return", 0) - fytd.get("bench_return", 0)), PCT2,
         "ahead of benchmark" if fytd.get("net_return", 0) >= fytd.get("bench_return", 0) else "trailing benchmark"),
        ("QUARTER TO DATE", qtd.get("net_return"), PCT2,
         f"vs benchmark {qtd.get('bench_return', 0):.2%}"),
        ("POLICY STATUS", None, None, ""),
        ("DAILY READ-THROUGH", tr_days[-1][1] if tr_days else None, "0.000%",
         f"coverage {tr_days[-1][2]:.1%} · {len(iss)} data issue(s)" if tr_days else ""),
    ]
    breaches = 0
    for cat, (lo, hi) in cfg["bands"].items():
        aw = d["alloc"].get(cat, {}).get("weight_actual")
        if aw is not None and not (lo <= aw <= hi):
            breaches += 1
    col = 1
    for label, val, fmt, sub in kpis:
        put(ws, 5, col, label, F_LBL, fill=FILL_TILE, border=True)
        if label == "POLICY STATUS":
            c = put(ws, 6, col, "0 breaches" if breaches == 0 else f"{breaches} BREACH(ES)", F_KPI,
                    fill=FILL_PASS if breaches == 0 else FILL_WARN, border=True)
            put(ws, 7, col, "vs IPS ranges", F_NOTE, fill=FILL_TILE, border=True)
        else:
            put(ws, 6, col, val, F_KPI, fmt=fmt, fill=FILL_TILE, border=True)
            put(ws, 7, col, sub, F_NOTE, fill=FILL_TILE, border=True)
        ws.column_dimensions[get_column_letter(col)].width = 24
        col += 1
    for extra_col in range(col, col + 6):
        ws.column_dimensions[get_column_letter(extra_col)].width = 13

    # charts from Data sheet
    lc = LineChart()
    lc.title = "Growth of $1 vs benchmark (demo fiscal year)"
    lc.style = 2
    lc.y_axis.numFmt = "0.00"
    lc.height, lc.width = 7.4, 12.6
    data = Reference(dws, min_col=4, max_col=5, min_row=3, max_row=3 + n_m)
    cats = Reference(dws, min_col=1, min_row=4, max_row=3 + n_m)
    lc.add_data(data, titles_from_data=True)
    lc.set_categories(cats)
    lc.series[0].graphicalProperties.line.solidFill = "3A6EA5"
    lc.series[0].graphicalProperties.line.width = 22000
    lc.series[1].graphicalProperties.line.solidFill = "C78F2E"
    lc.series[1].graphicalProperties.line.dashStyle = "dash"
    ws.add_chart(lc, "A9")

    bc = BarChart()
    bc.type = "col"
    bc.title = "Monthly net returns"
    bc.overlap = 100
    bc.height, bc.width = 7.4, 12.6
    bc.y_axis.numFmt = "0%"
    pos = Reference(dws, min_col=6, min_row=3, max_row=3 + n_m)
    neg = Reference(dws, min_col=7, min_row=3, max_row=3 + n_m)
    bc.add_data(pos, titles_from_data=True)
    bc.add_data(neg, titles_from_data=True)
    bc.set_categories(cats)
    bc.series[0].spPr = GraphicalProperties(solidFill="3A6EA5")
    bc.series[1].spPr = GraphicalProperties(solidFill="B4562A")
    bc.legend = None
    ws.add_chart(bc, "F9")

    tc = BarChart()
    tc.type = "col"
    tc.title = "Daily read-through (covered liquid proxies)"
    tc.overlap = 100
    tc.height, tc.width = 7.4, 12.6
    tc.y_axis.numFmt = "0.0%"
    n_t = len(tr_days)
    tpos = Reference(dws, min_col=4, min_row=rt0, max_row=rt0 + n_t)
    tneg = Reference(dws, min_col=5, min_row=rt0, max_row=rt0 + n_t)
    tcats = Reference(dws, min_col=1, min_row=rt0 + 1, max_row=rt0 + n_t)
    tc.add_data(tpos, titles_from_data=True)
    tc.add_data(tneg, titles_from_data=True)
    tc.set_categories(tcats)
    tc.series[0].spPr = GraphicalProperties(solidFill="3A6EA5")
    tc.series[1].spPr = GraphicalProperties(solidFill="B4562A")
    tc.legend = None
    ws.add_chart(tc, "K9")

    # allocation vs bands
    AR = 26
    put(ws, AR, 1, "Allocation vs IPS policy bands", F_H2)
    for j, h in enumerate(["Category", "Actual", "Target", "Band min", "Band max", "Status"]):
        put(ws, AR + 1, 1 + j, h, F_HDR, fill=FILL_HDR, border=True)
    rr = AR + 2
    for cat in CATS:
        a = d["alloc"].get(cat, {})
        aw, tw = a.get("weight_actual"), a.get("weight_target")
        band = cfg["bands"].get(cat)
        put(ws, rr, 1, CAT_LABEL[cat], border=True)
        put(ws, rr, 2, aw, fmt=PCT2, border=True)
        put(ws, rr, 3, tw, fmt=PCT1, border=True)
        put(ws, rr, 4, band[0] if band else None, fmt=PCT1, border=True)
        put(ws, rr, 5, band[1] if band else None, fmt=PCT1, border=True)
        status = ("WITHIN RANGE" if band and aw is not None and band[0] <= aw <= band[1]
                  else ("OUT OF RANGE" if band else "n/a"))
        put(ws, rr, 6, status, border=True,
            fill=FILL_PASS if status == "WITHIN RANGE" else (FILL_WARN if status == "OUT OF RANGE" else None))
        rr += 1

    # contribution + reconciliation
    CR = AR
    put(ws, CR, 8, "QTD contribution & reconciliation", F_H2)
    for j, h in enumerate(["Line", "Value"]):
        put(ws, CR + 1, 8 + j, h, F_HDR, fill=FILL_HDR, border=True)
    rr = CR + 2
    disp_sum = 0.0
    for cat, v in d["contrib"]:
        put(ws, rr, 8, CAT_LABEL[cat], border=True)
        put(ws, rr, 9, v, fmt=PCT2, border=True)
        disp_sum += round(v, 4)
        rr += 1
    ex = d["cextras"]
    resid = ex.get("residual", 0.0)
    chain = ex.get("return_chain_linked", 0.0)
    rounding = round(chain, 4) - disp_sum - round(resid, 4)
    for label, v in (("Displayed contributions", disp_sum), ("Compounding effect", resid),
                     ("Rounding adjustment", rounding), ("Chain-linked QTD return", chain)):
        put(ws, rr, 8, label, border=True)
        put(ws, rr, 9, v, fmt=PCT2, border=True)
        rr += 1
    ok = abs(resid) <= 0.001
    put(ws, rr, 8, "Status (10 bp tolerance)", border=True)
    put(ws, rr, 9, "PASS" if ok else "FAIL", border=True, fill=FILL_PASS if ok else FILL_WARN)

    # trend windows + issues
    TR = AR
    put(ws, TR, 11, "Proxy trend windows", F_H2)
    for j, h in enumerate(["Proxy", "Obs", "1d", "5d", "MTD", "σ20 (daily)"]):
        put(ws, TR + 1, 11 + j, h, F_HDR, fill=FILL_HDR, border=True)
    rr = TR + 2
    for pid in sorted(d["closes"]):
        t = trend(d["closes"][pid])
        put(ws, rr, 11, pid, border=True)
        put(ws, rr, 12, t["obs"], border=True)
        for j, key in enumerate(["d1", "d5", "mtd", "vol20"]):
            put(ws, rr, 13 + j, t[key], fmt=PCT2, border=True)
        rr += 1
    rr += 1
    put(ws, rr, 11, "Data issues", F_H2)
    rr += 1
    if iss:
        for msg in iss:
            put(ws, rr, 11, "! " + msg, F_TXT, fill=FILL_WARN, wrap=True)
            rr += 1
    else:
        put(ws, rr, 11, "None - all proxies current", fill=FILL_PASS)
        rr += 1
    checks = d["checks"]
    put(ws, rr + 1, 11,
        f"Workbook controls: {sum(1 for v in checks.values() if v == 'PASS')} PASS / "
        f"{sum(1 for v in checks.values() if v == 'WARN')} WARN / "
        f"{sum(1 for v in checks.values() if v == 'FAIL')} FAIL", F_NOTE)

# ---------------------------------------------------------------- policy sheet
ws = wb.create_sheet("Policy_IPS")
put(ws, 1, 1, "IPS policy reference (reported_public - quoted from the public Investment "
              "Policy Statements, restated June 12, 2024)", F_H2)
PENSION_TABLE = [
    ("Growth", 0.40, 0.48, 0.56, 0.505), ("Credit", 0.09, 0.13, 0.17, 0.12),
    ("Real Assets & Inflation Hedges", 0.11, 0.15, 0.19, 0.16),
    ("Risk Reduction & Mitigation", 0.16, 0.24, 0.32, 0.215),
]
OPEB_TABLE = [
    ("Growth", 0.35, 0.45, 0.55, 0.45), ("Credit", 0.11, 0.16, 0.21, 0.17),
    ("Real Assets & Inflation Hedges", 0.09, 0.13, 0.17, 0.165),
    ("Risk Reduction & Mitigation", 0.17, 0.26, 0.35, 0.215),
]
r = 3
for name, table, src in (("Pension", PENSION_TABLE, FUNDS["PENSION"]["ips"]),
                         ("OPEB Master Trust", OPEB_TABLE, FUNDS["OPEB"]["ips"])):
    put(ws, r, 1, f"{name} - category-level bands ({src[0]}, {src[1]})", F_H2)
    for j, h in enumerate(["Category", "Min", "Target", "Max", "1/2-step (eff 2024-07-01)"]):
        put(ws, r + 1, 1 + j, h, F_HDR, fill=FILL_HDR, border=True)
    for k, (label, lo, tgt, hi, hs) in enumerate(table):
        rr = r + 2 + k
        put(ws, rr, 1, label, border=True)
        for j, v in enumerate((lo, tgt, hi, hs)):
            put(ws, rr, 2 + j, v, fmt=PCT1, border=True)
    r += 8
put(ws, r, 1, "Bands are explicit min/target/max (Pension Cash is 1% with +2/-1 => 0-3%: never "
              "reconstruct from a +/- half-width). Sub-class detail and benchmark lags: Policy "
              "page of the web prototype.", F_NOTE, wrap=True)
ws.column_dimensions["A"].width = 30
for c_ in "BCDE":
    ws.column_dimensions[c_].width = 12

wb.save(OUT)
print(f"wrote {OUT}")
