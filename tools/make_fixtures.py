"""Generate deterministic synthetic fixtures under data/sample/ from the workbook export.

Reads the Excel-cached values of Export_Contract (the sole workbook->web interface) and writes:
  - data/sample/demofund_export_v1.csv        (valid, 338 records)
  - data/sample/invalid/*.csv                 (deliberately malformed variants for import tests)
  - data/sample/README.md                     (provenance and regeneration note)
"""
import csv
import datetime as dt
import io
import os

import openpyxl

import sys

ENTITY = (sys.argv[sys.argv.index("--entity") + 1] if "--entity" in sys.argv else "PENSION").upper()
SUFFIX = "" if ENTITY == "PENSION" else "_OPEB"
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WB = os.path.join(ROOT, "outputs", f"Portfolio_Analytics_Dashboard_Workbook_Prototype{SUFFIX}.xlsx")
OUT = os.path.join(ROOT, "data", "sample")
INV = os.path.join(OUT, "invalid")
os.makedirs(INV, exist_ok=True)

wb = openpyxl.load_workbook(WB, data_only=True)
ws = wb["Export_Contract"]
NCOL = 32  # schema 1.2: 29 base columns + entered_by / reviewed_by / review_status
hdr = [ws.cell(row=6, column=j + 1).value for j in range(NCOL)]
assert hdr[0] == "record_id" and hdr[-1] == "review_status", hdr

def fmt(v):
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, float):
        return format(v, ".10g")
    return str(v)

rows = []
for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
    if row[0].value is None:
        continue
    rows.append([fmt(row[j].value) for j in range(NCOL)])
assert len(rows) == 376, len(rows)  # 338 core + 20 schema-1.1 policy + 18 schema-1.3 recon/PM rows

def write(path, header, data):
    with io.open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(header)
        w.writerows(data)

VALID = os.path.join(
    OUT, "demofund_export_v1.csv" if ENTITY == "PENSION" else "demo_opeb_export_v1.csv"
)
write(VALID, hdr, rows)
print(f"wrote {VALID} ({len(rows)} records)")

if ENTITY != "PENSION":
    # invalid variants and the README derive from the Pension fixture only
    sys.exit(0)

# --- deliberately malformed variants (each begins from the valid data) -------------
i_class = hdr.index("classification")
i_val = hdr.index("value")
i_unit = hdr.index("unit")
i_qs = hdr.index("quality_status")
i_ver = hdr.index("schema_version")
i_pe = hdr.index("period_end")
i_ps = hdr.index("period_start")

def clone():
    return [list(r) for r in rows]

# 1. unsupported schema version
d = clone()
for r in d:
    r[i_ver] = "9.9.9"
write(os.path.join(INV, "bad_schema_version.csv"), hdr, d)

# 2. missing required column (classification dropped entirely)
h2 = [h for h in hdr if h != "classification"]
d = [[c for j, c in enumerate(r) if j != i_class] for r in rows]
write(os.path.join(INV, "missing_column.csv"), h2, d)

# 3. duplicate record ids + duplicate semantic keys (first data row repeated)
d = clone()
d.append(list(d[0]))
write(os.path.join(INV, "duplicate_records.csv"), hdr, d)

# 4. non-numeric value where a number is required
d = clone()
d[0][i_val] = "twelve"
write(os.path.join(INV, "bad_number.csv"), hdr, d)

# 5. whole-number percent (4.17 instead of 0.0417) — plausibility rejection
d = clone()
for r in d:
    if r[i_unit] == "%" and r[i_val] not in ("", None):
        r[i_val] = format(float(r[i_val]) * 100, ".10g")
write(os.path.join(INV, "whole_number_percent.csv"), hdr, d)

# 6. blank value not flagged missing
d = clone()
d[3][i_val] = ""
d[3][i_qs] = "ok"
write(os.path.join(INV, "blank_value_flagged_ok.csv"), hdr, d)

# 7. incoherent period (start after end) — mutate a period_return row
d = clone()
mutated = False
for r in d:
    if r[1] == "period_return" and r[i_ps] and r[i_pe] and r[i_ps] != r[i_pe]:
        r[i_ps], r[i_pe] = r[i_pe], r[i_ps]
        mutated = True
        break
assert mutated, "no period_return row found to mutate"
write(os.path.join(INV, "period_start_after_end.csv"), hdr, d)

# 8. invalid classification token
d = clone()
d[0][i_class] = "official_performance"
write(os.path.join(INV, "bad_classification.csv"), hdr, d)

# --- schema 1.2 provenance variants ------------------------------------------------
i_src = hdr.index("source_type")
i_entered = hdr.index("entered_by")
i_reviewed = hdr.index("reviewed_by")
i_rstatus = hdr.index("review_status")

# 9. user_import row without entered_by (V19)
d = clone()
d[0][i_src] = "user_import"
d[0][i_entered] = ""
write(os.path.join(INV, "user_import_no_entered_by.csv"), hdr, d)

# 10. published row without a reviewer (V20)
d = clone()
d[0][i_reviewed] = ""
assert d[0][i_rstatus] == "published"
write(os.path.join(INV, "published_no_reviewer.csv"), hdr, d)

# 11. invalid review_status token (V21)
d = clone()
d[0][i_rstatus] = "approved"
write(os.path.join(INV, "bad_review_status.csv"), hdr, d)

# 12. partial provenance header — reviewed_by column dropped (V02)
h12 = [h for h in hdr if h != "reviewed_by"]
d = [[c for j, c in enumerate(r) if j != i_reviewed] for r in rows]
write(os.path.join(INV, "partial_provenance_columns.csv"), h12, d)

print("wrote 12 invalid fixtures under", INV)

README = """# Sample data (synthetic)

`demofund_export_v1.csv` is the normalized export of the synthetic demonstration fund
`DEMOFUND` produced by the analyst workbook (`tools/build_workbook.py`, deterministic seed
20260630; exported by `tools/make_fixtures.py`). It contains **no actual portfolio data**;
the only `reported_public` rows are individually cited quotations from public LACERA
documents. Schema: 1.3.0 (29 base columns + `entered_by`/`reviewed_by`/`review_status`
provenance; actor labels are synthetic — `PA-ANALYST-1`, `PA-LEAD-1` — never real names).
See `docs/data-contract.md` and `docs/data-dictionary.md`.

`invalid/` contains deliberately malformed variants used to test the import validator
(`docs/import-validation-rules.md`). Each file name states its defect.

`demo_opeb_export_v1.csv` is the OPEB-entity equivalent (`DEMO-OPEB`), produced by the same
generator with `--entity OPEB` from its own Excel workbook. Both fixtures carry 16
`policy_target` and 4 `benchmark_definition` records (schema 1.1) quoting the public IPS band
structure, plus schema-1.3 rows: reconciliation pairs with tolerance-as-data (one deliberate
demo break) and a three-sleeve synthetic private-markets capital account (ratios are computed
by the app, never imported).

`demo_acfr_status_v1.csv` (entity `DEMO-ACFR`, generated by `tools/make_acfr_fixture.py`)
carries the ACFR section-readiness board: `acfr_section_status` history rows and
`acfr_artifact_link` rows with synthetic links/actors only.

Regenerate (QA steps require desktop Excel):
`python tools/build_workbook.py && python tools/qa_excel.py && python tools/make_fixtures.py`
then the same three commands with `--entity OPEB`.
"""
with io.open(os.path.join(OUT, "README.md"), "w", encoding="utf-8") as f:
    f.write(README)
print("wrote data/sample/README.md")
