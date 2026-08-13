"""Build outputs/Authoring_Template.xlsx — per-record-type authoring sheets that export into
the one flat contract CSV (schema 1.3, 32 columns).

Sheets:
  README        how to use; the three Monday-morning failure modes
  Enums         every closed token list, straight from the contract (dropdown sources)
  Settings      entity, as-of date, entered_by (V19), review_status
  Daily_Closes  the daily authoring grid (proxy id + close) -> feeds Export_Rows
  Allocation    staging grid with a live weight-sum check (V13 pre-flight)
  Returns       staging grid with a red flag on |value| > 0.60 (V10 pre-flight)
  Export_Rows   formula-assembled 32-column market_close rows; save this sheet as CSV

Enum lists mirror app/src/lib/contract/schema.ts — regenerate after any schema change.
"""
import os

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "outputs", "Authoring_Template.xlsx")

# ---- contract vocabulary (mirror of schema.ts) ------------------------------------
RECORD_TYPES = ["monthly_return", "monthly_benchmark_return", "period_return", "allocation",
                "contribution_qtd", "market_close", "public_reference", "check_result",
                "policy_target", "benchmark_definition", "recon_value", "tolerance_definition",
                "acfr_section_status", "acfr_artifact_link", "pm_commitment",
                "pm_capital_account"]
CLASSIFICATIONS = ["reported_public", "synthetic", "proxy_estimate", "calculated"]
PERIOD_TYPES = ["D", "M", "Q", "FY", "1M", "QTD", "FYTD", "1Y", "ITD"]
FREQUENCIES = ["Daily", "Monthly", "Quarterly", "Annual", "Ad Hoc"]
QUALITY = ["ok", "missing"]
BOOKS = ["IBOR", "ABOR", "n/a"]
RETURN_METHODS = ["TWR", "MWR", "n/a"]
GROSS_NET = ["gross", "net", "n/a"]
SOURCE_TYPES = ["workbook", "public_report", "synthetic_generator", "user_import"]
REVIEW_STATUSES = ["draft", "reviewed", "published", "n/a"]
CONTRACT_COLS = ["record_id", "record_type", "entity_id", "metric_id", "category_id", "value",
                 "unit", "currency", "scale", "as_of_date", "period_start", "period_end",
                 "period_type", "frequency", "classification", "source_type", "source_name",
                 "page_table", "provider", "retrieved_date", "book_of_record", "return_method",
                 "gross_net", "valuation_status", "benchmark_id", "method_id", "quality_status",
                 "note", "schema_version", "entered_by", "reviewed_by", "review_status"]

F_TITLE = Font(bold=True, size=14, color="1F3864")
F_H2 = Font(bold=True, size=11, color="1F3864")
F_HDR = Font(bold=True, size=10, color="FFFFFF")
F_TXT = Font(size=10)
F_NOTE = Font(italic=True, size=9, color="595959")
F_INPUT = Font(color="0000CC", size=10)
FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_INPUT = PatternFill("solid", fgColor="DDEBF7")
FILL_WARN = PatternFill("solid", fgColor="FCE4E4")
FILL_DISC = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

def hdr_row(ws, r, cols, start=1):
    for j, h in enumerate(cols):
        c = ws.cell(row=r, column=start + j, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.border = BORDER

# ---- README -----------------------------------------------------------------------
ws = wb.active
ws.title = "README"
ws["A1"] = "Contract authoring template — Fund Pulse prototype (schema 1.3)"
ws["A1"].font = F_TITLE
ws["A2"] = ("SYNTHETIC/DEMO WORKFLOW AID. Blue cells are inputs. This template stages rows for "
            "the one validated CSV contract; the dashboard's Import preflight remains the "
            "authority. Licensed market data: internal use only; never commit exports.")
ws["A2"].font = F_NOTE
ws["A2"].fill = FILL_DISC
notes = [
    ("Daily flow", "1) Set Settings (entity, as-of, your entered_by label). 2) Fill "
     "Daily_Closes with the day's proxy closes. 3) Open Export_Rows — it assembles valid "
     "market_close contract rows. 4) Save-As CSV (Export_Rows sheet only) and import in the "
     "app; append to the growing daily file to deepen every trend window."),
    ("Three Monday-morning mistakes", "Whole-number percents (4.17 for 4.17%) — V10 rejects; "
     "enter 0.0417. Two funds in one file — V17 rejects; one entity per file. Blank value "
     "without quality_status=missing — V08 rejects."),
    ("Staging sheets", "Allocation and Returns are pre-flight staging: the weight-sum cell "
     "mirrors V13 and the red flag mirrors V10 before you ever reach the validator."),
    ("Enums", "Every closed token list lives on the Enums sheet and feeds the dropdowns. "
     "Regenerate this template after any contract change (tools/make_authoring_template.py)."),
]
r = 4
for k, v in notes:
    ws.cell(row=r, column=1, value=k).font = F_H2
    c = ws.cell(row=r, column=2, value=v)
    c.font = F_TXT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 52
    r += 1
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 100

# ---- Enums ------------------------------------------------------------------------
ws = wb.create_sheet("Enums")
ws["A1"] = "Closed token lists (dropdown sources) — mirror of the app validator"
ws["A1"].font = F_TITLE
ENUM_COLS = [("record_type", RECORD_TYPES), ("classification", CLASSIFICATIONS),
             ("period_type", PERIOD_TYPES), ("frequency", FREQUENCIES),
             ("quality_status", QUALITY), ("book_of_record", BOOKS),
             ("return_method", RETURN_METHODS), ("gross_net", GROSS_NET),
             ("source_type", SOURCE_TYPES), ("review_status", REVIEW_STATUSES)]
for j, (name, values) in enumerate(ENUM_COLS):
    col = 1 + j
    c = ws.cell(row=3, column=col, value=name)
    c.font = F_HDR
    c.fill = FILL_HDR
    for i, v in enumerate(values):
        ws.cell(row=4 + i, column=col, value=v).font = F_TXT
    ws.column_dimensions[get_column_letter(col)].width = max(14, len(name) + 2)

def enum_ref(name, values):
    col = get_column_letter(1 + [e[0] for e in ENUM_COLS].index(name))
    return f"Enums!${col}$4:${col}${3 + len(values)}"

# ---- Settings ---------------------------------------------------------------------
ws = wb.create_sheet("Settings")
ws["A1"] = "File-level settings (blue = your input)"
ws["A1"].font = F_TITLE
settings = [("Entity (one per file, V17)", "DEMOFUND"),
            ("As-of date (YYYY-MM-DD)", "2026-07-01"),
            ("Your entered_by label (V19)", "PA-ANALYST-2"),
            ("review_status for new rows", "draft"),
            ("Starting record number", 9001)]
for i, (label, default) in enumerate(settings):
    ws.cell(row=3 + i, column=1, value=label).font = F_H2
    c = ws.cell(row=3 + i, column=2, value=default)
    c.font = F_INPUT
    c.fill = FILL_INPUT
    c.border = BORDER
dv = DataValidation(type="list", formula1=enum_ref("review_status", REVIEW_STATUSES),
                    allow_blank=False)
ws.add_data_validation(dv)
dv.add("B6")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 18
ws["A9"] = ("New rows default to review_status=draft: the dashboard shows a draft banner "
            "until a reviewer promotes them — that is the publish gate working.")
ws["A9"].font = F_NOTE

# ---- Daily_Closes -----------------------------------------------------------------
ws = wb.create_sheet("Daily_Closes")
ws["A1"] = "Daily closes (blue = the analyst's input). Feeds Export_Rows."
ws["A1"].font = F_TITLE
hdr_row(ws, 3, ["proxy_id", "close (index level)", "note (read-through mapping)"])
DEMO_PROXIES = ["DEMO-EQ-GLOBAL", "DEMO-EQ-SMALL", "DEMO-BOND-AGG", "DEMO-GOLD", "DEMO-OIL",
                "DEMO-USD"]
for i, p in enumerate(DEMO_PROXIES):
    ws.cell(row=4 + i, column=1, value=p).font = F_TXT
    c = ws.cell(row=4 + i, column=2)
    c.font = F_INPUT
    c.fill = FILL_INPUT
    c.border = BORDER
    c.number_format = "0.0000"
    ws.cell(row=4 + i, column=3).font = F_TXT
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 40

# ---- Allocation staging -----------------------------------------------------------
ws = wb.create_sheet("Allocation")
ws["A1"] = "Allocation staging — weight-sum pre-flight (mirrors V13)"
ws["A1"].font = F_TITLE
hdr_row(ws, 3, ["category_id", "weight_actual (decimal)", "emv $mm"])
CATS = ["GROWTH", "CREDIT", "RAIH", "RRM", "OVERLAY", "OTHER"]
for i, cat in enumerate(CATS):
    ws.cell(row=4 + i, column=1, value=cat).font = F_TXT
    for col in (2, 3):
        c = ws.cell(row=4 + i, column=col)
        c.font = F_INPUT
        c.fill = FILL_INPUT
        c.border = BORDER
        c.number_format = "0.0000" if col == 2 else "#,##0.0"
ws["A11"] = "Weight sum (must be 1.0000 ± 0.0001):"
ws["A11"].font = F_H2
ws["B11"] = "=SUM(B4:B9)"
ws["B11"].number_format = "0.0000"
ws["C11"] = '=IF(ABS(B11-1)<=0.0001,"OK","DOES NOT SUM TO 100% - V13 will reject")'
ws["C11"].font = F_H2
ws.conditional_formatting.add(
    "C11", CellIsRule(operator="notEqual", formula=['"OK"'], fill=FILL_WARN))
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 40

# ---- Returns staging --------------------------------------------------------------
ws = wb.create_sheet("Returns")
ws["A1"] = "Return staging — decimal pre-flight (mirrors V10: |return| must be ≤ 0.60)"
ws["A1"].font = F_TITLE
hdr_row(ws, 3, ["category_id", "return (DECIMAL: 0.0417 = 4.17%)"])
for i in range(8):
    a = ws.cell(row=4 + i, column=1)
    a.font = F_TXT
    c = ws.cell(row=4 + i, column=2)
    c.font = F_INPUT
    c.fill = FILL_INPUT
    c.border = BORDER
    c.number_format = "0.0000"
ws.conditional_formatting.add(
    "B4:B11",
    CellIsRule(operator="greaterThan", formula=["0.6"], fill=FILL_WARN))
ws.conditional_formatting.add(
    "B4:B11",
    CellIsRule(operator="lessThan", formula=["-0.6"], fill=FILL_WARN))
ws["A13"] = "Red fill = looks like a whole-number percent; V10 will reject it. Divide by 100."
ws["A13"].font = F_NOTE
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 30

# ---- Export_Rows ------------------------------------------------------------------
ws = wb.create_sheet("Export_Rows")
ws["A1"] = "Assembled market_close contract rows — Save-As CSV (this sheet only), then import"
ws["A1"].font = F_TITLE
hdr_row(ws, 3, CONTRACT_COLS)
for i in range(len(DEMO_PROXIES)):
    r = 4 + i
    src = f"Daily_Closes!A{4 + i}"
    close = f"Daily_Closes!B{4 + i}"
    note = f"Daily_Closes!C{4 + i}"
    vals = [
        f'=IF({close}="","",CONCATENATE("REC-",Settings!$B$7+{i}))',   # record_id
        f'=IF({close}="","","market_close")',                           # record_type
        f'=IF({close}="","",Settings!$B$3)',                            # entity_id
        f'=IF({close}="","","close")',                                  # metric_id
        f'=IF({close}="","",{src})',                                    # category_id
        f'=IF({close}="","",{close})',                                  # value
        f'=IF({close}="","","px")', f'=IF({close}="","","USD")',
        f'=IF({close}="","","1")',
        f'=IF({close}="","",TEXT(Settings!$B$4,"yyyy-mm-dd"))',         # as_of_date
        "", "", "",                                                       # ps, pe, pt
        f'=IF({close}="","","Daily")',
        f'=IF({close}="","","synthetic")',
        f'=IF({close}="","","user_import")',
        f'=IF({close}="","","Authoring_Template")',
        "",                                                               # page_table
        f'=IF({close}="","","analyst input")',
        f'=IF({close}="","",TEXT(Settings!$B$4,"yyyy-mm-dd"))',         # retrieved_date
        "", "", "",                                                       # book, rm, gn
        f'=IF({close}="","","final")',
        "", "",                                                           # bench, method
        f'=IF({close}="","","ok")',
        f'=IF({close}="","",{note})',
        f'=IF({close}="","","1.3.0")',
        f'=IF({close}="","",Settings!$B$5)',                            # entered_by (V19)
        "",                                                               # reviewed_by
        f'=IF({close}="","",Settings!$B$6)',                            # review_status
    ]
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=1 + j, value=v if v != "" else None)
        c.font = F_TXT
ws["A12"] = ("Rows appear as closes are entered. For licensed data run this workflow "
             "internally only. Settings!B7 spaces record ids away from the daily file's ids; "
             "adjust if appending to a file that already uses this range.")
ws["A12"].font = F_NOTE
for j in range(len(CONTRACT_COLS)):
    ws.column_dimensions[get_column_letter(1 + j)].width = 13

# Settings row numbers shifted: B3 entity, B4 as-of, B5 entered_by, B6 review_status, B7 start
wb.save(OUT)
print(f"saved {OUT}")
