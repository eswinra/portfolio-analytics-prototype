"""Stage 4 independent audit of the expanded workbook.

Reads the workbook file only (no reuse of build-time expectations): recomputes the full
calculation chain from the Inputs sheets and compares against Excel's cached values; audits
structure (hardcodes in calc cells, formula-pattern breaks, export consistency, enums, signs).
"""
import datetime as dt
import re
from collections import Counter

import openpyxl

import os
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WB = os.path.join(ROOT, "outputs", "Portfolio_Analytics_Dashboard_Workbook_Prototype.xlsx")
wbf = openpyxl.load_workbook(WB, data_only=False)
wbv = openpyxl.load_workbook(WB, data_only=True)

issues = []   # (severity, sheet, cell, text)
notes = []

def add(sev, sheet, cell, text):
    issues.append((sev, sheet, cell, text))

# ---------------------------------------------------------------- read inputs from file
ip = wbv["Inputs_Portfolio"]
CATS = ["GROWTH", "CREDIT", "RAIH", "RRM", "OVERLAY", "OTHER"]
bmv0, r6 = {}, 6
for k, cid in enumerate(CATS):
    assert ip.cell(row=6 + k, column=1).value == cid
    bmv0[cid] = ip.cell(row=6 + k, column=3).value

def find_grid(ws, header):  # locate section by its H2 label in column A
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if isinstance(c.value, str) and c.value.startswith(header):
            return c.row
    raise KeyError(header)

pr0 = find_grid(ip, "Monthly category net returns") + 2
br0 = find_grid(ip, "Monthly category benchmark net returns") + 2
tr0 = find_grid(ip, "End-of-month internal transfers") + 2
months, port_r, bench_r, trans = [], {c: [] for c in CATS}, {c: [] for c in CATS[:4]}, {c: [] for c in CATS}
for i in range(12):
    months.append(ip.cell(row=pr0 + i, column=1).value.date())
    for j, cid in enumerate(CATS):
        port_r[cid].append(ip.cell(row=pr0 + i, column=2 + j).value)
        trans[cid].append(ip.cell(row=tr0 + i, column=2 + j).value)
    for j, cid in enumerate(CATS[:4]):
        bench_r[cid].append(ip.cell(row=br0 + i, column=2 + j).value)

pt = wbv["Policy_Targets"]
as_of = pt["B4"].value.date()
hurdle = pt["B5"].value
eff1, eff2 = pt["B8"].value.date(), pt["D8"].value.date()
t1 = {cid: pt.cell(row=9 + k, column=2).value for k, cid in enumerate(CATS[:4])}
t2 = {cid: pt.cell(row=9 + k, column=4).value for k, cid in enumerate(CATS[:4])}

# ---------------------------------------------------------------- independent recompute
bmv = {c: [None] * 12 for c in CATS}
emv = {c: [None] * 12 for c in CATS}
for c in CATS:
    bmv[c][0] = bmv0[c]
for i in range(12):
    for c in CATS:
        emv[c][i] = bmv[c][i] * (1 + port_r[c][i]) + trans[c][i]
        if i < 11:
            bmv[c][i + 1] = emv[c][i]
bt = [sum(bmv[c][i] for c in CATS) for i in range(12)]
et = [sum(emv[c][i] for c in CATS) for i in range(12)]
w = {c: [bmv[c][i] / bt[i] for i in range(12)] for c in CATS}
tf = [sum(w[c][i] * port_r[c][i] for c in CATS) for i in range(12)]
gidx = []
g = 1.0
for x in tf:
    g *= 1 + x
    gidx.append(g)
btf = []
for i in range(12):
    mstart = months[i].replace(day=1)
    tw = t2 if mstart >= eff2 else t1
    btf.append(sum(tw[c] * bench_r[c][i] for c in CATS[:4]))
bidx = []
g = 1.0
for x in btf:
    g *= 1 + x
    bidx.append(g)
qtd, fytd = gidx[11] / gidx[8] - 1, gidx[11] - 1
bqtd, bfytd = bidx[11] / bidx[8] - 1, bidx[11] - 1
contrib = {c: sum(w[c][i] * port_r[c][i] for i in (9, 10, 11)) for c in CATS}
arith = sum(contrib.values())
resid = qtd - arith

# compare with Excel cached values
cr = wbv["Calc_Returns"]
TF_R0 = find_grid(wbf["Calc_Returns"], "Total fund monthly return") + 2
PER_R0 = find_grid(wbf["Calc_Returns"], "Period results") + 2
for i in range(12):
    got = cr.cell(row=TF_R0 + i, column=2).value
    if abs(got - tf[i]) > 1e-9:
        add("Critical", "Calc_Returns", f"B{TF_R0+i}", f"TF monthly {i}: excel {got} vs indep {tf[i]}")
per = [(cr.cell(row=PER_R0 + k, column=2).value, cr.cell(row=PER_R0 + k, column=3).value) for k in range(3)]
for label, got, exp in [("1M", per[0][0], tf[11]), ("QTD", per[1][0], qtd), ("FYTD", per[2][0], fytd),
                        ("bench 1M", per[0][1], btf[11]), ("bench QTD", per[1][1], bqtd),
                        ("bench FYTD", per[2][1], bfytd)]:
    if abs(got - exp) > 1e-9:
        add("Critical", "Calc_Returns", label, f"{label}: excel {got} vs indep {exp}")
cc = wbv["Calc_Contribution"]
for k, cid in enumerate(CATS):
    got = cc.cell(row=7 + k, column=11).value
    if abs(got - contrib[cid]) > 1e-9:
        add("Critical", "Calc_Contribution", f"K{7+k}", f"contribution {cid}: {got} vs {contrib[cid]}")
if abs(cc.cell(row=15, column=11).value - resid) > 1e-9:
    add("Critical", "Calc_Contribution", "K15", "residual mismatch")
if abs(resid) > 0.0010:
    add("Critical", "Calc_Contribution", "K16", f"residual {resid} exceeds tolerance")
ca = wbv["Calc_Allocation"]
for k, cid in enumerate(CATS):
    got = ca.cell(row=6 + k, column=3).value
    exp = emv[cid][11] / et[11]
    if abs(got - exp) > 1e-9:
        add("Critical", "Calc_Allocation", f"C{6+k}", f"alloc {cid}: {got} vs {exp}")
    tgt = ca.cell(row=6 + k, column=4).value
    exp_t = (t2 if as_of >= eff2 else t1).get(cid, 0)
    if abs((tgt or 0) - exp_t) > 1e-12:
        add("Critical", "Calc_Allocation", f"D{6+k}", f"target {cid}: {tgt} vs {exp_t}")
notes.append(f"independent recompute ties: TF QTD {qtd:.6f}, FYTD {fytd:.6f}, residual {resid:.6f}")

# transfers must net to zero
for i in range(12):
    s = sum(trans[c][i] for c in CATS)
    if abs(s) > 0.005:
        add("Critical", "Inputs_Portfolio", f"month {i}", f"transfers net {s}")

# ---------------------------------------------------------------- structural audit
ERR = re.compile(r"#(REF|VALUE|DIV/0|NAME|N/A|NUM|NULL)")
for ws in wbv.worksheets:
    if ws.sheet_state != "visible":
        add("Warning", ws.title, "-", "hidden sheet present")
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and ERR.search(c.value):
                add("Critical", ws.title, c.coordinate, f"error value {c.value}")

# hardcodes / formula-pattern breaks inside Calc_ sheets
for name in ("Calc_Returns", "Calc_Contribution", "Calc_Allocation", "Checks"):
    ws = wbf[name]
    wsv = wbv[name]
    for row in ws.iter_rows(min_row=5):
        for c in row:
            if c.value is None or c.column == 1:
                continue
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                continue
            # numeric constants allowed only in labeled input cells (blue): tolerance B4, targets
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                fill = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else None
                if fill != "00DDEBF7" and fill != "FFDDEBF7":
                    add("Warning", name, c.coordinate, f"numeric constant {v!r} in calc area (not marked input)")

# formula consistency along each 12-row grid column in Calc_Returns
ws = wbf["Calc_Returns"]
def shape(f):
    return re.sub(r"\d+", "#", f) if isinstance(f, str) else f
for top in range(1, ws.max_row):
    pass  # grids checked implicitly below
for grid_top, cols in ((TF_R0, [2, 3, 4, 5, 6]),):
    for col in cols:
        shapes = {shape(ws.cell(row=grid_top + i, column=col).value) for i in range(1, 12)}
        if len(shapes) > 1:
            add("Warning", "Calc_Returns", f"col {col}", f"inconsistent formulas: {shapes}")

# ---------------------------------------------------------------- export table audit
ex = wbv["Export_Contract"]
hdr = [ex.cell(row=6, column=j + 1).value for j in range(29)]
recs = []
for row in ex.iter_rows(min_row=7, max_row=ex.max_row):
    if row[0].value is None:
        continue
    recs.append({hdr[j]: row[j].value for j in range(29)})
notes.append(f"export records read: {len(recs)}")
ids = [r["record_id"] for r in recs]
if len(ids) != len(set(ids)):
    add("Critical", "Export_Contract", "-", "duplicate record_id")
CLASSES = {"reported_public", "synthetic", "proxy_estimate", "calculated"}
PTYPES = {"D", "M", "Q", "FY", "1M", "QTD", "FYTD", "1Y", "ITD"}
seen_keys = Counter()
for r in recs:
    rid = r["record_id"]
    if r["classification"] not in CLASSES:
        add("Critical", "Export_Contract", rid, f"bad classification {r['classification']!r}")
    if r["period_type"] not in PTYPES:
        add("Warning", "Export_Contract", rid, f"period_type {r['period_type']!r} not in enum")
    ps, pe = r["period_start"], r["period_end"]
    if ps and pe and ps > pe:
        add("Critical", "Export_Contract", rid, "period_start > period_end")
    if r["record_type"] == "market_close" and (r["value"] in (None, "")) and r["quality_status"] != "missing":
        add("Critical", "Export_Contract", rid, "blank value without missing flag")
    if r["value"] in (None, "") and r["quality_status"] == "ok" and r["record_type"] != "market_close":
        add("Critical", "Export_Contract", rid, "blank value flagged ok")
    if r["unit"] == "%" and isinstance(r["value"], (int, float)) and abs(r["value"]) > 1:
        add("Warning", "Export_Contract", rid, f"% value {r['value']} looks like whole-number percent")
    if r["classification"] == "reported_public" and r["record_type"] not in ("public_reference", "policy_target", "benchmark_definition"):
        add("Critical", "Export_Contract", rid, "reported_public outside public_reference")
    key = (r["record_type"], r["metric_id"], r["category_id"], str(r["as_of_date"]),
           str(r["period_start"]), str(r["period_end"]))
    seen_keys[key] += 1
dups = {k: n for k, n in seen_keys.items() if n > 1}
if dups:
    add("Critical", "Export_Contract", "-", f"duplicate semantic keys: {list(dups)[:3]}")
# scale/unit coherence
for r in recs:
    u, s = r["unit"], str(r["scale"])
    if u == "$mm" and s != "mm":
        add("Warning", "Export_Contract", r["record_id"], f"unit {u} but scale {s}")
    if u in ("$K", "$B") and s == "1":
        add("Warning", "Export_Contract", r["record_id"], f"unit {u} but scale '1' (fold into scale)")

# expected counts by record type
exp_counts = {"monthly_return": 84, "monthly_benchmark_return": 60, "period_return": 9,
              "allocation": 24, "contribution_qtd": 9, "market_close": 132,
              "public_reference": 8, "check_result": 12,
              "policy_target": 16, "benchmark_definition": 4}
got_counts = Counter(r["record_type"] for r in recs)
for k, v in exp_counts.items():
    if got_counts.get(k) != v:
        add("Critical", "Export_Contract", k, f"count {got_counts.get(k)} != {v}")

# check statuses in export match Checks sheet
chk = wbv["Checks"]
chk_status = {chk.cell(row=6 + i, column=1).value: chk.cell(row=6 + i, column=4).value for i in range(12)}
for r in recs:
    if r["record_type"] == "check_result" and chk_status.get(r["metric_id"]) != r["value"]:
        add("Critical", "Export_Contract", r["record_id"],
            f"check {r['metric_id']} export {r['value']} vs sheet {chk_status.get(r['metric_id'])}")

# validations point at Lists
for name, expect in (("Crosswalk", 1), ("QA_Checklist", 3)):
    dvs = wbf[name].data_validations.dataValidation
    if len(dvs) != expect:
        add("Warning", name, "-", f"{len(dvs)} validations, expected {expect}")
    for dv in dvs:
        if not str(dv.formula1).startswith("Lists!"):
            add("Warning", name, str(dv.sqref), f"validation source {dv.formula1!r}")

print("NOTES:")
for n in notes:
    print(" ", n)
print(f"\nISSUES ({len(issues)}):")
for sev, sheet, cell, text in sorted(issues, key=lambda x: {"Critical": 0, "Warning": 1, "Info": 2}[x[0]]):
    print(f"  [{sev}] {sheet}!{cell}: {text}")
if not issues:
    print("  none")
